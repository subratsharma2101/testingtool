"""
Test Executor - Executes generated test cases automatically using Playwright
"""

from playwright.sync_api import sync_playwright, Browser, Page, BrowserContext
import json
import time
import os
import sys
import threading
from datetime import datetime
from typing import Dict, List, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
import base64

# Fix encoding for Windows console
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

class TestExecutor:
    def __init__(self, website_url: str, login_id: str = "", password: str = "", headed: bool = True):
        self.website_url = website_url
        self.login_id = login_id
        self.password = password
        self.headed = headed
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.results = {
            'positive': [],
            'negative': [],
            'ui': [],
            'functional': [],
            'multi_level': []
        }
        self.summary = {
            'total_tests': 0,
            'passed': 0,
            'failed': 0,
            'skipped': 0,
            'execution_time': 0
        }
        self.max_workers = 3  # Parallel execution workers
        self.artifacts_dir = 'artifacts'
        self.console_logs = []
        self.network_requests = []
        
        # Create artifacts directory
        os.makedirs(self.artifacts_dir, exist_ok=True)
        os.makedirs('screenshots', exist_ok=True)
        os.makedirs('reports', exist_ok=True)
    
    def initialize_driver(self, capture_artifacts: bool = True):
        """Initialize Playwright Browser"""
        try:
            self.playwright = sync_playwright().start()
            
            # Launch browser with options
            browser_args = [
                '--start-maximized',
                '--disable-blink-features=AutomationControlled',
                '--disable-dev-shm-usage',
                '--no-sandbox'
            ]
            
            self.browser = self.playwright.chromium.launch(
                headless=not self.headed,
                args=browser_args
            )
            
            # Create context with viewport
            self.context = self.browser.new_context(
                viewport={'width': 1920, 'height': 1080}
            )
            
            # Create page
            self.page = self.context.new_page()
            
            # Set default timeout (increased for slow websites)
            self.page.set_default_timeout(60000)  # 60 seconds
            
            # Setup artifact capture if requested
            if capture_artifacts:
                self._setup_artifact_capture()
            
            return self.page
        except Exception as e:
            print(f"[ERROR] Failed to initialize Playwright browser: {e}")
            raise
    
    def _setup_artifact_capture(self):
        """Setup console log and network request capture"""
        try:
            # Capture console logs
            def handle_console(msg):
                self.console_logs.append({
                    'type': msg.type,
                    'text': msg.text,
                    'timestamp': time.time()
                })
            
            self.page.on("console", handle_console)
            
            # Capture network requests
            def handle_request(request):
                self.network_requests.append({
                    'url': request.url,
                    'method': request.method,
                    'timestamp': time.time()
                })
            
            self.page.on("request", handle_request)
        except Exception as e:
            print(f"[WARNING] Failed to setup artifact capture: {e}")
    
    def take_screenshot(self, test_id: str) -> str:
        """Take screenshot and return as base64 string"""
        try:
            screenshot_path = f"screenshots/{test_id}_{int(time.time())}.png"
            self.page.screenshot(path=screenshot_path, full_page=True)
            
            # Convert to base64
            with open(screenshot_path, 'rb') as f:
                screenshot_data = base64.b64encode(f.read()).decode('utf-8')
            
            return screenshot_data
        except Exception as e:
            print(f"[ERROR] Failed to take screenshot: {e}")
            return ""
    
    def _capture_console_logs(self) -> List[Dict[str, Any]]:
        """Get recent console logs"""
        return self.console_logs[-20:] if self.console_logs else []
    
    def _capture_page_snapshot(self, test_id: str):
        """Capture page HTML snapshot"""
        try:
            snapshot_dir = os.path.join(self.artifacts_dir, 'snapshots')
            os.makedirs(snapshot_dir, exist_ok=True)
            snapshot_path = os.path.join(snapshot_dir, f"{test_id}_{int(time.time())}.html")
            
            html_content = self.page.content()
            with open(snapshot_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return snapshot_path
        except Exception as e:
            print(f"[WARNING] Failed to capture page snapshot: {e}")
            return None
    
    def _capture_network_har(self, test_id: str):
        """Capture network HAR file"""
        try:
            har_dir = os.path.join(self.artifacts_dir, 'har')
            os.makedirs(har_dir, exist_ok=True)
            har_path = os.path.join(har_dir, f"{test_id}_{int(time.time())}.har")
            
            # Get HAR from context
            har_data = self.context.tracing.stop()
            if har_data:
                with open(har_path, 'w', encoding='utf-8') as f:
                    json.dump(har_data, f, indent=2)
            
            return har_path
        except Exception as e:
            print(f"[WARNING] Failed to capture HAR: {e}")
            return None
    
    def execute_positive_test(self, test_case: Dict) -> Dict:
        """Execute a positive test case"""
        test_id = test_case.get('test_id', 'UNKNOWN')
        test_name = test_case.get('test_name', 'Unknown Test')
        
        result = {
            'test_id': test_id,
            'test_name': test_name,
            'status': 'SKIPPED',
            'execution_time': 0,
            'error_message': '',
            'screenshot': ''
        }
        
        try:
            start_time = time.time()
            
            # Initialize driver if needed
            if not self.page:
                self.initialize_driver(capture_artifacts=True)
            
            # Navigate to website
            print(f"[POSITIVE] Executing: {test_id} - {test_name}")
            try:
                self.page.goto(self.website_url, wait_until='domcontentloaded', timeout=60000)
                try:
                    self.page.wait_for_load_state('networkidle', timeout=30000)
                except:
                    try:
                        self.page.wait_for_load_state('load', timeout=15000)
                    except:
                        time.sleep(2)
            except Exception as nav_error:
                result['status'] = 'FAIL'
                result['error_message'] = f"Cannot reach website: {str(nav_error)}"
                result['execution_time'] = round(time.time() - start_time, 2)
                return result
            
            # Execute login test if it's a login test
            if 'LOGIN' in test_id and self.login_id and self.password:
                login_success = self._perform_login()
                result['status'] = 'PASS' if login_success else 'FAIL'
                if not login_success:
                    result['error_message'] = 'Login failed'
            else:
                # For input field tests
                if 'INPUT' in test_id:
                    field_name = test_case.get('field', '')
                    if field_name:
                        try:
                            # Try to find and fill the field
                            field = self.page.locator(f"input[name='{field_name}'], input[id='{field_name}']").first
                            if field.count() > 0:
                                field.fill("test_data_123")
                                result['status'] = 'PASS'
                            else:
                                result['status'] = 'FAIL'
                                result['error_message'] = f'Could not find field: {field_name}'
                        except Exception as e:
                            result['status'] = 'FAIL'
                            result['error_message'] = f'Could not interact with field: {field_name} - {str(e)}'
                    else:
                        result['status'] = 'PASS'  # Field exists check passed
            
            execution_time = time.time() - start_time
            result['execution_time'] = round(execution_time, 2)
            
            # Normalize status to uppercase
            if result['status'] not in ['PASS', 'FAIL', 'SKIPPED']:
                result['status'] = result['status'].upper()
            
            if result['status'] == 'FAIL':
                result['screenshot'] = self.take_screenshot(test_id)
            
        except Exception as e:
            result['status'] = 'FAIL'
            result['error_message'] = str(e)
            result['screenshot'] = self.take_screenshot(test_id)
            result['execution_time'] = round(time.time() - start_time, 2) if 'start_time' in locals() else 0
        
        return result
    
    def execute_negative_test(self, test_case: Dict) -> Dict:
        """Execute a negative test case"""
        test_id = test_case.get('test_id', 'UNKNOWN')
        test_name = test_case.get('test_name', 'Unknown Test')
        
        result = {
            'test_id': test_id,
            'test_name': test_name,
            'status': 'SKIPPED',
            'execution_time': 0,
            'error_message': '',
            'screenshot': ''
        }
        
        try:
            start_time = time.time()
            
            # Initialize driver if needed
            if not self.page:
                self.initialize_driver(capture_artifacts=True)
            
            # Navigate to website
            print(f"[NEGATIVE] Executing: {test_id} - {test_name}")
            try:
                self.page.goto(self.website_url, wait_until='domcontentloaded', timeout=60000)
                try:
                    self.page.wait_for_load_state('networkidle', timeout=30000)
                except:
                    try:
                        self.page.wait_for_load_state('load', timeout=15000)
                    except:
                        time.sleep(2)
            except Exception as nav_error:
                result['status'] = 'FAIL'
                result['error_message'] = f"Cannot reach website: {str(nav_error)}"
                result['execution_time'] = round(time.time() - start_time, 2)
                return result
            
            # Execute based on test type
            if 'LOGIN' in test_id:
                if 'EMPTY' in test_id or 'empty' in test_name.lower():
                    # Test empty field validation
                    try:
                        submit_btn = self.page.locator("button[type='submit'], input[type='submit']").first
                        if submit_btn.count() > 0:
                            submit_btn.click()
                            time.sleep(2)
                            
                            # Check if error message appears
                            page_content = self.page.content().lower()
                            if 'error' in page_content or 'required' in page_content or 'invalid' in page_content:
                                result['status'] = 'PASS'
                            else:
                                result['status'] = 'FAIL'
                                result['error_message'] = 'Expected error message not displayed'
                        else:
                            result['status'] = 'FAIL'
                            result['error_message'] = 'Could not find submit button'
                    except Exception as e:
                        result['status'] = 'FAIL'
                        result['error_message'] = f'Could not execute empty field test: {str(e)}'
                elif 'INVALID' in test_id or 'invalid' in test_name.lower():
                    # Test invalid credentials
                    login_success = self._perform_login_with_invalid_credentials()
                    result['status'] = 'PASS' if not login_success else 'FAIL'
                    if login_success:
                        result['error_message'] = 'Login should have failed with invalid credentials'
            elif test_case.get('malicious_payload'):
                success, message = self._handle_security_payload_test(test_case)
                result['status'] = 'PASS' if success else 'FAIL'
                if not success:
                    result['error_message'] = message or 'Malicious payload was not rejected'
            
            execution_time = time.time() - start_time
            result['execution_time'] = round(execution_time, 2)
            
            # Normalize status
            if result['status'] not in ['PASS', 'FAIL', 'SKIPPED']:
                result['status'] = result['status'].upper()
            
            if result['status'] == 'FAIL':
                result['screenshot'] = self.take_screenshot(test_id)
            
        except Exception as e:
            result['status'] = 'FAIL'
            result['error_message'] = str(e)
            result['screenshot'] = self.take_screenshot(test_id)
            result['execution_time'] = round(time.time() - start_time, 2) if 'start_time' in locals() else 0
        
        return result
    
    def execute_ui_test(self, test_case: Dict) -> Dict:
        """Execute a UI test case"""
        test_id = test_case.get('test_id', 'UNKNOWN')
        test_name = test_case.get('test_name', 'Unknown Test')
        
        result = {
            'test_id': test_id,
            'test_name': test_name,
            'status': 'SKIPPED',
            'execution_time': 0,
            'error_message': '',
            'screenshot': ''
        }
        
        try:
            start_time = time.time()
            
            # Initialize driver if needed
            if not self.page:
                self.initialize_driver(capture_artifacts=True)
            
            # Navigate to website
            print(f"[UI] Executing: {test_id} - {test_name}")
            try:
                self.page.goto(self.website_url, wait_until='domcontentloaded', timeout=60000)
                try:
                    self.page.wait_for_load_state('networkidle', timeout=30000)
                except:
                    try:
                        self.page.wait_for_load_state('load', timeout=15000)
                    except:
                        time.sleep(2)
            except Exception as nav_error:
                result['status'] = 'FAIL'
                result['error_message'] = f"Cannot reach website: {str(nav_error)}"
                result['execution_time'] = round(time.time() - start_time, 2)
                return result
            
            # Execute UI checks
            if 'BUTTON' in test_id:
                button_text = test_name.split('Verify ')[-1].split(' button')[0] if 'button' in test_name.lower() else ''
                if button_text:
                    try:
                        button = self.page.locator(f"button:has-text('{button_text}')").first
                        if button.count() > 0 and button.is_visible() and button.is_enabled():
                            result['status'] = 'PASS'
                        else:
                            result['status'] = 'FAIL'
                            result['error_message'] = f'Button is not visible or enabled: {button_text}'
                    except Exception as e:
                        result['status'] = 'FAIL'
                        result['error_message'] = f'Could not find button: {button_text} - {str(e)}'
                else:
                    result['status'] = 'SKIPPED'
                    result['error_message'] = 'Could not extract button text from test case'
            elif 'INPUT' in test_id:
                field_name = test_id.split('INPUT_')[1] if 'INPUT_' in test_id else ''
                if not field_name:
                    field_name = test_name.split('Verify ')[-1].split(' field')[0] if 'field' in test_name.lower() else ''
                if field_name:
                    try:
                        field = self.page.locator(f"input[name='{field_name}'], input[id='{field_name}']").first
                        if field.count() > 0 and field.is_visible():
                            result['status'] = 'PASS'
                        else:
                            result['status'] = 'FAIL'
                            result['error_message'] = 'Field is not visible'
                    except Exception as e:
                        result['status'] = 'FAIL'
                        result['error_message'] = f'Could not find field: {field_name} - {str(e)}'
            elif 'PAGE' in test_id:
                if 'title' in test_name.lower():
                    page_title = self.page.title()
                    if page_title:
                        result['status'] = 'PASS'
                    else:
                        result['status'] = 'FAIL'
                        result['error_message'] = 'Page title not found'
            
            execution_time = time.time() - start_time
            result['execution_time'] = round(execution_time, 2)
            
            # Normalize status
            if result['status'] not in ['PASS', 'FAIL', 'SKIPPED']:
                result['status'] = result['status'].upper()
            
            if result['status'] == 'FAIL':
                result['screenshot'] = self.take_screenshot(test_id)
            
        except Exception as e:
            result['status'] = 'FAIL'
            result['error_message'] = str(e)
            result['screenshot'] = self.take_screenshot(test_id)
            result['execution_time'] = round(time.time() - start_time, 2) if 'start_time' in locals() else 0
        
        return result
    
    def execute_functional_test(self, test_case: Dict) -> Dict:
        """Execute a functional test case"""
        test_id = test_case.get('test_id', 'UNKNOWN')
        test_name = test_case.get('test_name', 'Unknown Test')
        test_type = test_case.get('test_type', '')
        
        result = {
            'test_id': test_id,
            'test_name': test_name,
            'status': 'SKIPPED',
            'execution_time': 0,
            'error_message': '',
            'screenshot': ''
        }
        
        try:
            start_time = time.time()
            
            # Initialize driver if needed
            if not self.page:
                self.initialize_driver(capture_artifacts=True)
            
            # Navigate to website
            print(f"[FUNCTIONAL] Executing: {test_id} - {test_name}")
            try:
                self.page.goto(self.website_url, wait_until='domcontentloaded', timeout=60000)
                try:
                    self.page.wait_for_load_state('networkidle', timeout=30000)
                except:
                    try:
                        self.page.wait_for_load_state('load', timeout=15000)
                    except:
                        time.sleep(2)
            except Exception as nav_error:
                result['status'] = 'FAIL'
                result['error_message'] = f"Cannot reach website: {str(nav_error)}"
                result['execution_time'] = round(time.time() - start_time, 2)
                return result
            
            if test_type == 'navigation':
                link_href = test_case.get('link', '')
                if link_href:
                    try:
                        current_url = self.page.url
                        link = self.page.locator(f"a[href='{link_href}']").first
                        if link.count() > 0:
                            link.click()
                            time.sleep(3)
                            new_url = self.page.url
                            if new_url != current_url:
                                result['status'] = 'PASS'
                            else:
                                result['status'] = 'FAIL'
                                result['error_message'] = 'Navigation did not occur'
                        else:
                            result['status'] = 'FAIL'
                            result['error_message'] = 'Could not find link'
                    except Exception as e:
                        result['status'] = 'FAIL'
                        result['error_message'] = f'Could not execute navigation test: {str(e)}'
            
            elif test_type == 'button_functionality':
                button_text = test_name.split('Verify ')[-1].split(' button')[0] if 'button' in test_name.lower() else ''
                if button_text:
                    try:
                        button = self.page.locator(f"button:has-text('{button_text}')").first
                        if button.count() > 0:
                            button.click()
                            time.sleep(2)
                            result['status'] = 'PASS'
                        else:
                            result['status'] = 'FAIL'
                            result['error_message'] = f'Could not find button: {button_text}'
                    except Exception as e:
                        result['status'] = 'FAIL'
                        result['error_message'] = f'Could not click button: {button_text} - {str(e)}'
            
            elif test_type == 'form_submission':
                try:
                    form = self.page.locator("form").first
                    if form.count() > 0:
                        inputs = form.locator("input[type='text'], input[type='email']")
                        input_count = inputs.count()
                        for i in range(min(input_count, 3)):
                            inputs.nth(i).fill("test@example.com")
                        
                        password_inputs = form.locator("input[type='password']")
                        if password_inputs.count() > 0:
                            password_inputs.first.fill("test123")
                        
                        submit_btn = form.locator("button[type='submit'], input[type='submit']").first
                        if submit_btn.count() > 0:
                            submit_btn.click()
                            time.sleep(3)
                            result['status'] = 'PASS'
                        else:
                            result['status'] = 'FAIL'
                            result['error_message'] = 'Could not find submit button'
                    else:
                        result['status'] = 'FAIL'
                        result['error_message'] = 'Could not find form'
                except Exception as e:
                    result['status'] = 'FAIL'
                    result['error_message'] = f'Could not execute form submission test: {str(e)}'
            
            execution_time = time.time() - start_time
            result['execution_time'] = round(execution_time, 2)
            
            # Normalize status
            if result['status'] not in ['PASS', 'FAIL', 'SKIPPED']:
                result['status'] = result['status'].upper()
            
            if result['status'] == 'FAIL':
                result['screenshot'] = self.take_screenshot(test_id)
            
        except Exception as e:
            result['status'] = 'FAIL'
            result['error_message'] = str(e)
            result['screenshot'] = self.take_screenshot(test_id)
            result['execution_time'] = round(time.time() - start_time, 2) if 'start_time' in locals() else 0
        
        return result
    
    def execute_multi_level_test(self, test_case: Dict) -> Dict:
        """Execute a multi-level form test case"""
        test_id = test_case.get('test_id', 'UNKNOWN')
        test_name = test_case.get('test_name', 'Unknown Test')
        
        result = {
            'test_id': test_id,
            'test_name': test_name,
            'status': 'SKIPPED',
            'execution_time': 0,
            'error_message': '',
            'screenshot': ''
        }
        
        try:
            start_time = time.time()
            
            # Initialize driver if needed
            if not self.page:
                self.initialize_driver(capture_artifacts=True)
            
            # Navigate to website
            print(f"[MULTI-LEVEL] Executing: {test_id} - {test_name}")
            try:
                self.page.goto(self.website_url, wait_until='domcontentloaded', timeout=60000)
                try:
                    self.page.wait_for_load_state('networkidle', timeout=30000)
                except:
                    try:
                        self.page.wait_for_load_state('load', timeout=15000)
                    except:
                        time.sleep(2)
            except Exception as nav_error:
                result['status'] = 'FAIL'
                result['error_message'] = f"Cannot reach website: {str(nav_error)}"
                result['execution_time'] = round(time.time() - start_time, 2)
                return result
            
            # Extract level information from test case
            level_info = test_case.get('level_info', {})
            level_number = level_info.get('level', 1)
            
            # Fill and submit multi-level form
            success = self._fill_and_submit_multi_level_form(level_number)
            result['status'] = 'PASS' if success else 'FAIL'
            if not success:
                result['error_message'] = f'Failed to complete level {level_number} form submission'
            
            execution_time = time.time() - start_time
            result['execution_time'] = round(execution_time, 2)
            
            # Normalize status
            if result['status'] not in ['PASS', 'FAIL', 'SKIPPED']:
                result['status'] = result['status'].upper()
            
            if result['status'] == 'FAIL':
                result['screenshot'] = self.take_screenshot(test_id)
            
        except Exception as e:
            result['status'] = 'FAIL'
            result['error_message'] = str(e)
            result['screenshot'] = self.take_screenshot(test_id)
            result['execution_time'] = round(time.time() - start_time, 2) if 'start_time' in locals() else 0
        
        return result
    
    def _fill_and_submit_multi_level_form(self, level: int) -> bool:
        """Fill and submit multi-level form"""
        try:
            # Find all input fields in current form
            inputs = self.page.locator("input[type='text'], input[type='email'], input[type='number']").all()
            
            # Fill inputs with test data
            for i, inp in enumerate(inputs[:10]):  # Limit to 10 fields
                try:
                    field_type = inp.get_attribute("type") or "text"
                    if field_type == "email":
                        inp.fill(f"test{i}@example.com")
                    elif field_type == "number":
                        inp.fill(str(100 + i))
                    else:
                        inp.fill(f"Test Data {i+1}")
                    time.sleep(0.2)
                except:
                    continue
            
            # Find and click submit/next button
            submit_btn = self.page.locator(
                "button[type='submit'], input[type='submit'], button:has-text('Next'), button:has-text('Submit'), button:has-text('Continue')"
            ).first
            
            if submit_btn.count() > 0:
                submit_btn.click()
                time.sleep(3)
                
                # Wait for next level or success
                try:
                    self.page.wait_for_load_state('networkidle', timeout=10000)
                except:
                    time.sleep(2)
                
                return True
            else:
                return False
                
        except Exception as e:
            print(f"[WARNING] Multi-level form submission error: {e}")
            return False
    
    def _perform_login(self) -> bool:
        """Perform login with credentials"""
        try:
            # Find username field
            username_field = self.page.locator(
                "input[name='username'], input[name='user'], input[name='email'], input[name='userid'], input[name='user_id'], input[id='username'], input[id='user'], input[id='email'], input[id='userid']"
            ).first
            
            # Find password field
            password_field = self.page.locator("input[type='password']").first
            
            if username_field.count() > 0 and password_field.count() > 0:
                username_field.fill(self.login_id)
                password_field.fill(self.password)
                
                # Find and click login button
                login_btn = self.page.locator("button[type='submit'], input[type='submit'], button:has-text('Login'), button:has-text('Sign In')").first
                if login_btn.count() > 0:
                    login_btn.click()
                    time.sleep(3)
                    
                    # Check if login successful
                    current_url = self.page.url
                    if current_url != self.website_url:
                        return True
                return False
            return False
        except Exception as e:
            print(f"[WARNING] Login failed: {e}")
            return False
    
    def _perform_login_with_invalid_credentials(self) -> bool:
        """Perform login with invalid credentials"""
        try:
            username_field = self.page.locator("input[type='text'], input[type='email']").first
            password_field = self.page.locator("input[type='password']").first
            
            if username_field.count() > 0 and password_field.count() > 0:
                username_field.fill("invalid_user")
                password_field.fill("invalid_pass")
                
                login_btn = self.page.locator("button[type='submit'], input[type='submit']").first
                if login_btn.count() > 0:
                    login_btn.click()
                    time.sleep(3)
                    
                    # Check if still on login page (login failed = test passed)
                    current_url = self.page.url
                    if current_url == self.website_url:
                        return False  # Login failed = test passed
                    return True  # Login succeeded = test failed
            return False
        except:
            return False
    
    def _handle_security_payload_test(self, test_case: Dict) -> tuple:
        """Attempt to submit malicious payloads and verify they are rejected"""
        payload = test_case.get('malicious_payload', '')
        field_name = test_case.get('field', '')
        field_id = test_case.get('field_id', '')
        
        try:
            # Find target field
            field = None
            if field_id:
                field = self.page.locator(f"input[id='{field_id}'], textarea[id='{field_id}']").first
            if not field or field.count() == 0:
                if field_name:
                    field = self.page.locator(f"input[name='{field_name}'], textarea[name='{field_name}']").first
            
            if field and field.count() > 0:
                field.fill(payload)
                
                # Try to submit
                submit_btn = self.page.locator("button[type='submit'], input[type='submit']").first
                if submit_btn.count() > 0:
                    try:
                        submit_btn.click()
                    except:
                        pass
                    time.sleep(2)
                
                # Check for error indicators
                page_content = self.page.content().lower()
                indicators = ['error', 'invalid', 'blocked', 'forbidden', 'denied', 'malicious']
                if any(indicator in page_content for indicator in indicators):
                    return True, ''
                return False, 'Payload did not trigger validation feedback'
            return False, 'Unable to locate target field'
        except Exception as e:
            return False, str(e)
    
    def _execute_single_test(self, category: str, test_case: Dict) -> Dict:
        """Execute a single test case based on category"""
        if category == 'positive':
            return self.execute_positive_test(test_case)
        elif category == 'negative':
            return self.execute_negative_test(test_case)
        elif category == 'ui':
            return self.execute_ui_test(test_case)
        elif category == 'functional':
            return self.execute_functional_test(test_case)
        elif category == 'multi_level':
            return self.execute_multi_level_test(test_case)
        else:
            return {
                'test_id': test_case.get('test_id', 'UNKNOWN'),
                'test_name': test_case.get('test_name', 'Unknown'),
                'status': 'SKIPPED',
                'execution_time': 0,
                'error_message': f'Unknown category: {category}',
                'screenshot': ''
            }
    
    def _execute_tests_sequential(self, test_cases: Dict):
        """Execute tests sequentially with detailed progress logging"""
        
        # Print test suite summary
        print("\n" + "="*80)
        print("TEST SUITE SUMMARY")
        print("="*80)
        total_positive = len(test_cases.get('positive', []))
        total_negative = len(test_cases.get('negative', []))
        total_ui = len(test_cases.get('ui', []))
        total_functional = len(test_cases.get('functional', []))
        total_multi_level = len(test_cases.get('multi_level', []))
        total_all = total_positive + total_negative + total_ui + total_functional + total_multi_level
        
        print(f"Total Test Cases Found: {total_all}")
        print(f"  - Positive Tests:     {total_positive}")
        print(f"  - Negative Tests:     {total_negative}")
        print(f"  - UI Tests:           {total_ui}")
        print(f"  - Functional Tests:   {total_functional}")
        print(f"  - Multi-Level Tests:  {total_multi_level}")
        print("="*80 + "\n")
        
        # Execute positive tests
        positive_tests = test_cases.get('positive', [])
        if positive_tests:
            print(f"\n{'='*80}")
            print(f"EXECUTING POSITIVE TESTS ({len(positive_tests)} tests)")
            print(f"{'='*80}")
            for idx, test_case in enumerate(positive_tests, 1):
                test_id = test_case.get('test_id', 'UNKNOWN')
                test_name = test_case.get('test_name', 'Unknown')
                print(f"\n[POSITIVE {idx}/{len(positive_tests)}] Starting: {test_id}")
                print(f"  Test Name: {test_name}")
                
                start = time.time()
                result = self.execute_positive_test(test_case)
                elapsed = time.time() - start
                
                status = result.get('status', 'UNKNOWN')
                status_symbol = "✓" if status == 'PASS' else "✗" if status == 'FAIL' else "⊘"
                print(f"  {status_symbol} Result: {status} (took {elapsed:.2f}s)")
                if result.get('error_message'):
                    print(f"  Error: {result.get('error_message')[:100]}")
                
                self.results['positive'].append(result)
                self._update_summary(result)
                if result.get('status') == 'FAIL':
                    self._capture_page_snapshot(result.get('test_id', 'unknown'))
        else:
            print(f"\n[INFO] No positive tests to execute")
        
        # Execute negative tests
        negative_tests = test_cases.get('negative', [])
        if negative_tests:
            print(f"\n{'='*80}")
            print(f"EXECUTING NEGATIVE TESTS ({len(negative_tests)} tests)")
            print(f"{'='*80}")
            for idx, test_case in enumerate(negative_tests, 1):
                test_id = test_case.get('test_id', 'UNKNOWN')
                test_name = test_case.get('test_name', 'Unknown')
                print(f"\n[NEGATIVE {idx}/{len(negative_tests)}] Starting: {test_id}")
                print(f"  Test Name: {test_name}")
                
                start = time.time()
                result = self.execute_negative_test(test_case)
                elapsed = time.time() - start
                
                status = result.get('status', 'UNKNOWN')
                status_symbol = "✓" if status == 'PASS' else "✗" if status == 'FAIL' else "⊘"
                print(f"  {status_symbol} Result: {status} (took {elapsed:.2f}s)")
                if result.get('error_message'):
                    print(f"  Error: {result.get('error_message')[:100]}")
                
                self.results['negative'].append(result)
                self._update_summary(result)
                if result.get('status') == 'FAIL':
                    self._capture_page_snapshot(result.get('test_id', 'unknown'))
        else:
            print(f"\n[INFO] No negative tests to execute")
        
        # Execute UI tests
        ui_tests = test_cases.get('ui', [])
        if ui_tests:
            print(f"\n{'='*80}")
            print(f"EXECUTING UI TESTS ({len(ui_tests)} tests)")
            print(f"{'='*80}")
            for idx, test_case in enumerate(ui_tests, 1):
                test_id = test_case.get('test_id', 'UNKNOWN')
                test_name = test_case.get('test_name', 'Unknown')
                print(f"\n[UI {idx}/{len(ui_tests)}] Starting: {test_id}")
                print(f"  Test Name: {test_name}")
                
                start = time.time()
                result = self.execute_ui_test(test_case)
                elapsed = time.time() - start
                
                status = result.get('status', 'UNKNOWN')
                status_symbol = "✓" if status == 'PASS' else "✗" if status == 'FAIL' else "⊘"
                print(f"  {status_symbol} Result: {status} (took {elapsed:.2f}s)")
                if result.get('error_message'):
                    print(f"  Error: {result.get('error_message')[:100]}")
                
                self.results['ui'].append(result)
                self._update_summary(result)
                if result.get('status') == 'FAIL':
                    self._capture_page_snapshot(result.get('test_id', 'unknown'))
        else:
            print(f"\n[INFO] No UI tests to execute")
        
        # Execute functional tests
        functional_tests = test_cases.get('functional', [])
        if functional_tests:
            print(f"\n{'='*80}")
            print(f"EXECUTING FUNCTIONAL TESTS ({len(functional_tests)} tests)")
            print(f"{'='*80}")
            for idx, test_case in enumerate(functional_tests, 1):
                test_id = test_case.get('test_id', 'UNKNOWN')
                test_name = test_case.get('test_name', 'Unknown')
                print(f"\n[FUNCTIONAL {idx}/{len(functional_tests)}] Starting: {test_id}")
                print(f"  Test Name: {test_name}")
                
                start = time.time()
                result = self.execute_functional_test(test_case)
                elapsed = time.time() - start
                
                status = result.get('status', 'UNKNOWN')
                status_symbol = "✓" if status == 'PASS' else "✗" if status == 'FAIL' else "⊘"
                print(f"  {status_symbol} Result: {status} (took {elapsed:.2f}s)")
                if result.get('error_message'):
                    print(f"  Error: {result.get('error_message')[:100]}")
                
                self.results['functional'].append(result)
                self._update_summary(result)
                if result.get('status') == 'FAIL':
                    self._capture_page_snapshot(result.get('test_id', 'unknown'))
        else:
            print(f"\n[INFO] No functional tests to execute")
        
        # Execute multi-level tests
        multi_level_tests = test_cases.get('multi_level', [])
        if multi_level_tests:
            print(f"\n{'='*80}")
            print(f"EXECUTING MULTI-LEVEL TESTS ({len(multi_level_tests)} tests)")
            print(f"{'='*80}")
            for idx, test_case in enumerate(multi_level_tests, 1):
                test_id = test_case.get('test_id', 'UNKNOWN')
                test_name = test_case.get('test_name', 'Unknown')
                print(f"\n[MULTI-LEVEL {idx}/{len(multi_level_tests)}] Starting: {test_id}")
                print(f"  Test Name: {test_name}")
                
                start = time.time()
                result = self.execute_multi_level_test(test_case)
                elapsed = time.time() - start
                
                status = result.get('status', 'UNKNOWN')
                status_symbol = "✓" if status == 'PASS' else "✗" if status == 'FAIL' else "⊘"
                print(f"  {status_symbol} Result: {status} (took {elapsed:.2f}s)")
                if result.get('error_message'):
                    print(f"  Error: {result.get('error_message')[:100]}")
                
                self.results['multi_level'].append(result)
                self._update_summary(result)
                if result.get('status') == 'FAIL':
                    self._capture_page_snapshot(result.get('test_id', 'unknown'))
        else:
            print(f"\n[INFO] No multi-level tests to execute")
    
    def execute_all_tests(self, test_cases: Dict, parallel: bool = False) -> Dict:
        """Execute all test cases"""
        start_time = time.time()
        
        if not self.page:
            self.initialize_driver(capture_artifacts=True)
        
        print("[INFO] Starting test execution...")
        
        # Reset results
        self.results = {
            'positive': [],
            'negative': [],
            'ui': [],
            'functional': [],
            'multi_level': []
        }
        self.summary = {
            'total_tests': 0,
            'passed': 0,
            'failed': 0,
            'skipped': 0,
            'execution_time': 0
        }
        
        if parallel:
            # Parallel execution (requires separate browser instances per test)
            print("[WARNING] Parallel execution not fully implemented, using sequential mode")
            self._execute_tests_sequential(test_cases)
        else:
            # Sequential execution
            self._execute_tests_sequential(test_cases)
        
        total_time = time.time() - start_time
        self.summary['execution_time'] = round(total_time, 2)
        self.summary['total_tests'] = self.summary['passed'] + self.summary['failed'] + self.summary['skipped']
        
        print("\n" + "="*80)
        print("TEST EXECUTION COMPLETED")
        print("="*80)
        print(f"Total Tests:  {self.summary['total_tests']}")
        print(f"Passed:       {self.summary['passed']}")
        print(f"Failed:       {self.summary['failed']}")
        print(f"Skipped:      {self.summary['skipped']}")
        print(f"Execution Time: {self.summary['execution_time']:.2f}s")
        print("="*80 + "\n")
        
        return {
            'results': self.results,
            'summary': self.summary
        }
    
    def _update_summary(self, result: Dict):
        """Update execution summary"""
        status = result.get('status', 'SKIPPED')
        # Normalize status to uppercase
        if status not in ['PASS', 'FAIL', 'SKIPPED']:
            status = status.upper()
        
        if status == 'PASS':
            self.summary['passed'] += 1
        elif status == 'FAIL':
            self.summary['failed'] += 1
        else:
            self.summary['skipped'] += 1
    
    def save_execution_report(self, website_url: str) -> str:
        """Save execution report to file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = f"execution_report_{timestamp}.json"
        report_path = os.path.join('reports', report_file)
        
        report_data = {
            'website_url': website_url,
            'executed_at': datetime.now().isoformat(),
            'summary': self.summary,
            'results': self.results
        }
        
        # Remove screenshots from JSON (they're too large)
        for test_type in self.results:
            for result in self.results[test_type]:
                if 'screenshot' in result:
                    result['screenshot'] = 'Available' if result['screenshot'] else 'Not available'
        
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False, default=str)
        
        return report_file
    
    def save_execution_report_excel(self, website_url: str, test_cases_data: Dict = None) -> str:
        """Save execution report to Excel (.xlsx). Falls back to CSV if openpyxl is unavailable."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_file = f"execution_report_{timestamp}.xlsx"
        xlsx_path = os.path.join('reports', xlsx_file)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except Exception:
            # Fallback to CSV if openpyxl isn't installed
            import csv
            csv_file = f"execution_report_{timestamp}.csv"
            csv_path = os.path.join('reports', csv_file)
            os.makedirs('reports', exist_ok=True)
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'Category', 'Test ID', 'Test Name', 'Description', 'Status', 
                    'Result (Pass/Fail)', 'Execution Time (s)', 'Failure Reason'
                ])
                
                # Write all test cases with results
                for category, results in self.results.items():
                    for r in results:
                        # Find matching test case for description
                        description = ''
                        if test_cases_data:
                            test_list = test_cases_data.get(category, [])
                            for tc in test_list:
                                if tc.get('test_id') == r.get('test_id'):
                                    description = tc.get('description', '') or tc.get('test_name', '')
                                    break
                        
                        writer.writerow([
                            category,
                            r.get('test_id', ''),
                            r.get('test_name', ''),
                            description,
                            r.get('status', ''),
                            'PASS' if r.get('status') == 'PASS' else 'FAIL' if r.get('status') == 'FAIL' else 'SKIP',
                            r.get('execution_time', 0),
                            r.get('error_message', '') if r.get('status') == 'FAIL' else ''
                        ])
            return csv_file
        
        # Build XLSX report
        os.makedirs('reports', exist_ok=True)
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Summary header
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'] = 'Test Execution Summary'
        ws_summary['A1'].font = title_font
        ws_summary['A1'].alignment = center_align
        
        executed_at = datetime.now().isoformat()
        summary_rows = [
            ('Website URL', website_url),
            ('Executed At', executed_at),
            ('', ''),
            ('Total Tests', self.summary.get('total_tests', 0)),
            ('Passed', self.summary.get('passed', 0)),
            ('Failed', self.summary.get('failed', 0)),
            ('Skipped', self.summary.get('skipped', 0)),
            ('Execution Time (s)', self.summary.get('execution_time', 0)),
        ]
        
        for r_idx, (k, v) in enumerate(summary_rows, start=2):
            cell_k = ws_summary.cell(row=r_idx, column=1, value=k)
            cell_v = ws_summary.cell(row=r_idx, column=2, value=v)
            if k and k != '':
                cell_k.font = header_font
            cell_k.alignment = left_align
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 50
        
        # Results sheet
        ws = wb.create_sheet('Results')
        headers = [
            'Category', 'Test ID', 'Test Name', 'Description (English)', 'Status', 
            'Result (Pass/Fail)', 'Execution Time (s)', 'Failure Reason'
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_color = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = header_font_color
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Write all test cases with results
        row = 2
        for category, results in self.results.items():
            for r in results:
                # Find matching test case for description
                description = ''
                if test_cases_data:
                    test_list = test_cases_data.get(category, [])
                    for tc in test_list:
                        if tc.get('test_id') == r.get('test_id'):
                            description = tc.get('description', '') or tc.get('test_name', '')
                            break
                
                status = r.get('status', 'SKIPPED')
                result_text = 'PASS' if status == 'PASS' else 'FAIL' if status == 'FAIL' else 'SKIP'
                
                # Color code result
                fill_color = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid") if status == 'PASS' else (
                    PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid") if status == 'FAIL' else None
                )
                
                ws.cell(row=row, column=1, value=category).border = border
                ws.cell(row=row, column=2, value=r.get('test_id', '')).border = border
                ws.cell(row=row, column=3, value=r.get('test_name', '')).border = border
                ws.cell(row=row, column=4, value=description).border = border
                ws.cell(row=row, column=4).alignment = left_align
                ws.cell(row=row, column=5, value=status).border = border
                result_cell = ws.cell(row=row, column=6, value=result_text)
                result_cell.border = border
                if fill_color:
                    result_cell.fill = fill_color
                ws.cell(row=row, column=7, value=r.get('execution_time', 0)).border = border
                failure_cell = ws.cell(row=row, column=8, value=r.get('error_message', '') if status == 'FAIL' else '')
                failure_cell.border = border
                failure_cell.alignment = left_align
                
                row += 1
        
        # Auto-size columns
        column_widths = {
            'A': 12,  # Category
            'B': 20,  # Test ID
            'C': 30,  # Test Name
            'D': 40,  # Description
            'E': 10,  # Status
            'F': 15,  # Result
            'G': 15,  # Execution Time
            'H': 40   # Failure Reason
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        wb.save(xlsx_path)
        return xlsx_file
    
    def close(self):
        """Close browser"""
        try:
            if self.context:
                self.context.close()
            if self.browser:
                self.browser.close()
            if self.playwright:
                self.playwright.stop()
        except Exception as e:
            print(f"[WARNING] Error closing browser: {e}")
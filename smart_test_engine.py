from playwright.sync_api import sync_playwright, Browser, Page, BrowserContext
import json
import time
import os
import sys
import re
from datetime import datetime
from typing import Dict, List, Any
from urllib.parse import urljoin

class SmartTestEngine:
    def __init__(self, website_url: str, login_id: str, password: str, headed: bool = True, otp_value: str = None):
        self.website_url = website_url
        self.login_id = login_id
        self.password = password
        self.headed = headed
        self.otp_value = otp_value
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.detected_elements = {}
        self.test_cases = {
            'positive': [],
            'negative': [],
            'ui': [],
            'functional': [],
            'workflow': [],
            'edge_case': []
        }
        self.test_priorities = {}  # Track priorities
        self.workflow_context = []  # Track multi-step workflows
    
    def initialize_driver(self):
        """Initialize Playwright Browser"""
        try:
            self.playwright = sync_playwright().start()
            
            # Launch browser with options
            browser_args = [
                '--start-maximized',
                '--disable-blink-features=AutomationControlled',
                '--disable-dev-shm-usage',
                '--no-sandbox'
            ]
            
            self.browser = self.playwright.chromium.launch(
                headless=not self.headed,
                args=browser_args
            )
            
            # Create context with viewport
            self.context = self.browser.new_context(
                viewport={'width': 1920, 'height': 1080}
            )
            
            # Create page
            self.page = self.context.new_page()
            
            # Set default timeout (increased for slow websites)
            self.page.set_default_timeout(60000)  # 60 seconds
            
            return self.page
        except Exception as e:
            print(f"[ERROR] Failed to initialize Playwright browser: {e}")
            raise
    
    def analyze_website(self):
        """Analyze website and detect all elements"""
        print("[INFO] Analyzing website structure...")
        if not self.page:
            self.initialize_driver()
        
        try:
            # Try to navigate with timeout
            print(f"[INFO] Navigating to: {self.website_url}")
            self.page.goto(self.website_url, wait_until='domcontentloaded', timeout=60000)
            
            # Wait for page to load with timeout
            try:
                self.page.wait_for_load_state('networkidle', timeout=30000)
            except Exception as e:
                print(f"[WARNING] Network idle timeout, trying load state: {e}")
                try:
                    self.page.wait_for_load_state('load', timeout=15000)
                except:
                    print("[WARNING] Load state timeout, continuing anyway...")
                    time.sleep(2)  # Give it a small delay
            
        except Exception as e:
            error_msg = str(e)
            print(f"[ERROR] Failed to navigate to website: {error_msg}")
            raise Exception(f"Cannot reach website '{self.website_url}'. Please check: 1) URL is correct, 2) Website is accessible, 3) Internet connection is working. Error: {error_msg}")
        
        detected = {
            'forms': [],
            'input_fields': [],
            'buttons': [],
            'links': [],
            'images': [],
            'dropdowns': [],
            'checkboxes': [],
            'radio_buttons': [],
            'textareas': [],
            'iframes': [],
            'tables': [],
            'page_title': self.page.title(),
            'current_url': self.page.url
        }
        
        # Detect Input Fields
        try:
            inputs = self.page.locator("input").all()
            for inp in inputs:
                try:
                    input_type = (inp.get_attribute("type") or "text").lower()
                    input_name = inp.get_attribute("name") or inp.get_attribute("id") or "unnamed"
                    placeholder = inp.get_attribute("placeholder") or ""
                    field_meta = {
                        'type': input_type,
                        'name': input_name,
                        'id': inp.get_attribute("id"),
                        'placeholder': placeholder,
                        'required': inp.get_attribute("required") is not None,
                        'class': inp.get_attribute("class") or ""
                    }
                    detected['input_fields'].append(field_meta)
                    
                    if input_type == 'checkbox':
                        detected['checkboxes'].append({
                            'name': input_name,
                            'id': field_meta['id'],
                            'label': inp.get_attribute("aria-label") or placeholder or input_name
                        })
                    elif input_type == 'radio':
                        detected['radio_buttons'].append({
                            'name': input_name,
                            'id': field_meta['id'],
                            'label': inp.get_attribute("aria-label") or placeholder or input_name
                        })
                except:
                    continue
        except Exception as e:
            print(f"Error detecting inputs: {e}")

        # Detect Textareas
        try:
            textareas = self.page.locator("textarea").all()
            for area in textareas:
                try:
                    area_name = area.get_attribute("name") or area.get_attribute("id") or "textarea"
                    field_meta = {
                        'type': 'textarea',
                        'name': area_name,
                        'id': area.get_attribute("id"),
                        'placeholder': area.get_attribute("placeholder") or "",
                        'required': area.get_attribute("required") is not None,
                        'class': area.get_attribute("class") or ""
                    }
                    detected['textareas'].append(field_meta)
                    detected['input_fields'].append(field_meta)
                except:
                    continue
        except Exception as e:
            print(f"Error detecting textareas: {e}")
        
        # Detect Buttons
        try:
            buttons = self.page.locator("button").all()
            buttons += self.page.locator("input[type='submit']").all()
            buttons += self.page.locator("input[type='button']").all()
            buttons += self.page.locator("a[class*='button'], a[class*='btn']").all()
            
            seen_buttons = set()
            for btn in buttons:
                try:
                    btn_text = btn.inner_text().strip() or btn.get_attribute("value") or btn.get_attribute("aria-label") or "Button"
                    btn_id = btn.get_attribute("id") or ""
                    
                    if btn_id in seen_buttons:
                        continue
                    seen_buttons.add(btn_id)
                    
                    tag_name = btn.evaluate("el => el.tagName.toLowerCase()")
                    detected['buttons'].append({
                        'text': btn_text,
                        'id': btn_id,
                        'class': btn.get_attribute("class") or "",
                        'type': btn.get_attribute("type") or "button",
                        'tag': tag_name
                    })
                except:
                    continue
        except Exception as e:
            print(f"Error detecting buttons: {e}")
        
        # Detect Links
        try:
            links = self.page.locator("a").all()[:30]
            seen_links = set()
            for link in links:
                try:
                    href = link.get_attribute("href")
                    link_text = link.inner_text().strip()
                    if href and (link_text or href not in seen_links):
                        seen_links.add(href)
                        detected['links'].append({
                            'text': link_text or href,
                            'href': href
                        })
                except:
                    continue
        except Exception as e:
            print(f"Error detecting links: {e}")
        
        # Detect Forms
        try:
            forms = self.page.locator("form").all()
            for form in forms:
                try:
                    detected['forms'].append({
                        'action': form.get_attribute("action") or "",
                        'method': form.get_attribute("method") or "get",
                        'id': form.get_attribute("id")
                    })
                except:
                    continue
        except Exception as e:
            print(f"Error detecting forms: {e}")
        
        # Detect Dropdowns
        try:
            selects = self.page.locator("select").all()
            for select in selects:
                try:
                    detected['dropdowns'].append({
                        'name': select.get_attribute("name") or select.get_attribute("id"),
                        'id': select.get_attribute("id")
                    })
                except:
                    continue
        except Exception as e:
            print(f"Error detecting dropdowns: {e}")

        # Detect Iframes
        try:
            frames = self.page.locator("iframe").all()[:10]
            for idx, frame in enumerate(frames):
                try:
                    detected['iframes'].append({
                        'id': frame.get_attribute("id") or f"iframe_{idx + 1}",
                        'name': frame.get_attribute("name"),
                        'src': frame.get_attribute("src")
                    })
                except:
                    continue
        except Exception as e:
            print(f"Error detecting iframes: {e}")

        # Detect Tables
        try:
            tables = self.page.locator("table").all()[:10]
            for idx, table in enumerate(tables):
                try:
                    headers = [th.inner_text().strip() for th in table.locator("th").all() if th.inner_text().strip()]
                    rows = table.locator("tr").all()
                    detected['tables'].append({
                        'id': table.get_attribute("id") or f"table_{idx + 1}",
                        'headers': headers[:8],
                        'row_count': len(rows)
                    })
                except:
                    continue
        except Exception as e:
            print(f"Error detecting tables: {e}")
        
        # Detect School ERP Modules
        detected['modules'] = self.detect_erp_modules()
        
        # Detect Dashboard Widgets
        detected['dashboard_widgets'] = self.detect_dashboard_widgets()
        
        # Enhanced Table Analysis for ERP
        detected['table_details'] = self.analyze_erp_tables(detected.get('tables', []))
        
        # Detect Multi-Level Forms (Level 1, 2, 3, etc.)
        detected['form_levels'] = self._detect_form_levels()
        
        self.detected_elements = detected
        return detected
    
    def detect_erp_modules(self) -> List[Dict[str, Any]]:
        """Detect School ERP modules from navigation menu"""
        modules = []
        erp_module_keywords = {
            'student': ['student', 'admission', 'enrollment', 'pupil', 'learner'],
            'teacher': ['teacher', 'faculty', 'staff', 'instructor'],
            'academic': ['course', 'subject', 'class', 'section', 'syllabus', 'curriculum'],
            'attendance': ['attendance', 'present', 'absent', 'leave'],
            'examination': ['exam', 'test', 'assessment', 'result', 'grade', 'marks'],
            'finance': ['fee', 'payment', 'expense', 'financial', 'billing', 'invoice'],
            'library': ['library', 'book', 'issue', 'return'],
            'hostel': ['hostel', 'room', 'boarding'],
            'transport': ['transport', 'bus', 'vehicle', 'route'],
            'report': ['report', 'analytics', 'dashboard', 'statistics']
        }
        
        try:
            # Check navigation menus (nav, sidebar, menu)
            nav_selectors = [
                "nav a",
                ".sidebar a",
                ".menu a",
                "[class*='nav'] a",
                "[class*='menu'] a",
                "[class*='sidebar'] a",
                ".navbar a"
            ]
            
            all_links = []
            for selector in nav_selectors:
                try:
                    links = self.page.locator(selector).all()
                    for link in links:
                        try:
                            text = link.inner_text().strip().lower()
                            href = link.get_attribute("href") or ""
                            if text and len(text) > 1:
                                all_links.append({
                                    'text': text,
                                    'href': href,
                                    'id': link.get_attribute("id") or ""
                                })
                        except:
                            continue
                except:
                    continue
            
            # Detect modules based on keywords
            detected_modules = {}
            for link in all_links:
                link_text = link['text']
                for module_type, keywords in erp_module_keywords.items():
                    if any(keyword in link_text for keyword in keywords):
                        if module_type not in detected_modules:
                            detected_modules[module_type] = {
                                'name': module_type.title(),
                                'links': [],
                                'count': 0
                            }
                        detected_modules[module_type]['links'].append({
                            'text': link['text'],
                            'href': link['href'],
                            'id': link['id']
                        })
                        detected_modules[module_type]['count'] += 1
            
            # Convert to list
            modules = list(detected_modules.values())
            
        except Exception as e:
            print(f"[WARNING] Error detecting ERP modules: {e}")
        
        return modules
    
    def detect_dashboard_widgets(self) -> List[Dict[str, Any]]:
        """Detect dashboard widgets (cards, stats, charts)"""
        widgets = []
        
        try:
            # Detect widget-like elements (cards, panels, stats)
            widget_selectors = [
                ".card",
                ".widget",
                ".panel",
                "[class*='card']",
                "[class*='widget']",
                "[class*='stat']",
                "[class*='dashboard']"
            ]
            
            seen_widgets = set()
            for selector in widget_selectors:
                try:
                    elements = self.page.locator(selector).all()[:20]
                    for element in elements:
                        try:
                            element_id = element.get_attribute("id") or ""
                            element_class = element.get_attribute("class") or ""
                            text = element.inner_text().strip()
                            
                            # Skip if too small or duplicate
                            if len(text) < 5 or element_id in seen_widgets:
                                continue
                            
                            seen_widgets.add(element_id)
                            
                            widgets.append({
                                'id': element_id or f"widget_{len(widgets) + 1}",
                                'class': element_class,
                                'title': text[:50] if text else '',
                                'type': self._classify_widget_type(element_class, text)
                            })
                        except:
                            continue
                except:
                    continue
            
        except Exception as e:
            print(f"[WARNING] Error detecting dashboard widgets: {e}")
        
        return widgets
    
    def _classify_widget_type(self, element_class: str, text: str) -> str:
        """Classify widget type based on class and content"""
        class_lower = element_class.lower()
        text_lower = text.lower()
        
        if 'chart' in class_lower or 'graph' in class_lower:
            return 'chart'
        elif 'stat' in class_lower or any(keyword in text_lower for keyword in ['total', 'count', 'number']):
            return 'statistics'
        elif 'table' in class_lower:
            return 'table'
        elif 'list' in class_lower:
            return 'list'
        else:
            return 'card'
    
    def analyze_erp_tables(self, tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Enhanced table analysis for School ERP"""
        table_details = []
        
        try:
            for table_info in tables:
                table_id = table_info.get('id', '')
                headers = table_info.get('headers', [])
                
                # Try to locate the table
                try:
                    if table_id and table_id.startswith('table_'):
                        table_locator = self.page.locator("table").first
                    else:
                        table_locator = self.page.locator(f"table#{table_id}").first
                    
                    if not table_locator.count():
                        table_locator = self.page.locator("table").first
                    
                    # Analyze table type based on headers
                    table_type = self._classify_erp_table(headers)
                    
                    # Get row count and sample data
                    rows = table_locator.locator("tbody tr, tr").all()
                    row_count = len(rows)
                    
                    # Get pagination info if exists
                    pagination = self._detect_table_pagination(table_locator)
                    
                    # Detect action buttons in table
                    action_buttons = self._detect_table_actions(table_locator)
                    
                    table_details.append({
                        'id': table_id,
                        'headers': headers,
                        'row_count': row_count,
                        'type': table_type,
                        'pagination': pagination,
                        'actions': action_buttons,
                        'has_search': self._has_table_search(table_locator),
                        'has_filter': self._has_table_filter(table_locator),
                        'has_export': self._has_table_export(table_locator)
                    })
                    
                except Exception as e:
                    print(f"[WARNING] Error analyzing table {table_id}: {e}")
                    continue
                    
        except Exception as e:
            print(f"[WARNING] Error in ERP table analysis: {e}")
        
        return table_details
    
    def _resolve_url(self, href: str) -> str:
        """Resolve relative URL to absolute URL"""
        if href.startswith('http'):
            return href
        elif href.startswith('/'):
            return urljoin(self.website_url, href)
        else:
            return urljoin(self.page.url, href)
    
    def _detect_form_levels(self) -> List[Dict[str, Any]]:
        """Detect different form levels or routes on the website"""
        form_levels = []
        
        try:
            # Look for navigation links that might lead to different form levels
            level_patterns = [
                r'level[\s_-]?1',
                r'level[\s_-]?2', 
                r'level[\s_-]?3',
                r'level[\s_-]?4',
                r'level[\s_-]?5',
                r'form[\s_-]?1',
                r'form[\s_-]?2',
                r'form[\s_-]?3',
                r'/level/1',
                r'/level/2',
                r'/level/3',
                r'/form/1',
                r'/form/2',
                r'/form/3'
            ]
            
            # Get all links
            links = self.page.locator("a").all()
            level_links = []
            
            for link in links:
                try:
                    href = link.get_attribute("href") or ""
                    link_text = link.inner_text().strip().lower()
                    
                    # Check if link contains level/form pattern
                    for pattern in level_patterns:
                        if re.search(pattern, href, re.IGNORECASE) or re.search(pattern, link_text, re.IGNORECASE):
                            full_url = self._resolve_url(href)
                            # Avoid duplicates
                            if full_url not in [l.get('full_url', '') for l in level_links]:
                                level_links.append({
                                    'text': link.inner_text().strip(),
                                    'href': href,
                                    'full_url': full_url
                                })
                            break
                except:
                    continue
            
            # If level links found, analyze each level
            if level_links:
                print(f"[INFO] Found {len(level_links)} form level links. Analyzing each level...")
                original_url = self.page.url
                
                for idx, level_link in enumerate(level_links[:5]):  # Max 5 levels
                    try:
                        level_url = level_link['full_url']
                        print(f"[INFO] Analyzing Level {idx + 1}: {level_url}")
                        
                        # Navigate to level
                        try:
                            print(f"[INFO] Navigating to level: {level_url}")
                            self.page.goto(level_url, wait_until='domcontentloaded', timeout=60000)
                            
                            # Wait for page load with fallback
                            try:
                                self.page.wait_for_load_state('networkidle', timeout=30000)
                            except:
                                try:
                                    self.page.wait_for_load_state('load', timeout=15000)
                                except:
                                    print("[WARNING] Level page load timeout, continuing...")
                                    time.sleep(2)
                        except Exception as e:
                            print(f"[WARNING] Failed to navigate to level {idx + 1}: {e}")
                            continue
                        
                        time.sleep(1)  # Small delay for page to stabilize
                        
                        # Detect forms on this level
                        forms = self.page.locator("form").all()
                        inputs = self.page.locator("input, textarea, select").all()
                        
                        level_data = {
                            'level_number': idx + 1,
                            'level_name': level_link['text'] or f'Level {idx + 1}',
                            'url': level_url,
                            'form_count': len(forms),
                            'input_count': len(inputs),
                            'forms': []
                        }
                        
                        # Analyze each form on this level
                        for form_idx, form in enumerate(forms):
                            try:
                                form_inputs = form.locator("input, textarea, select").all()
                                form_data = {
                                    'form_id': form.get_attribute("id") or f"form_{form_idx + 1}",
                                    'action': form.get_attribute("action") or "",
                                    'method': form.get_attribute("method") or "get",
                                    'input_fields': []
                                }
                                
                                for inp in form_inputs:
                                    try:
                                        tag_name = inp.evaluate("el => el.tagName.toLowerCase()")
                                        field_data = {
                                            'type': tag_name,
                                            'name': inp.get_attribute("name") or inp.get_attribute("id") or "",
                                            'id': inp.get_attribute("id"),
                                            'placeholder': inp.get_attribute("placeholder") or "",
                                            'required': inp.get_attribute("required") is not None,
                                            'label': ''
                                        }
                                        
                                        # Get label if exists
                                        try:
                                            field_id = inp.get_attribute("id")
                                            if field_id:
                                                label = self.page.locator(f"label[for='{field_id}']").first
                                                if label.count():
                                                    field_data['label'] = label.inner_text().strip()
                                        except:
                                            pass
                                        
                                        if tag_name == 'input':
                                            field_data['input_type'] = inp.get_attribute("type") or "text"
                                        elif tag_name == 'select':
                                            options = inp.locator("option").all()
                                            field_data['options'] = [opt.inner_text().strip() for opt in options[:10] if opt.inner_text().strip()]
                                        
                                        form_data['input_fields'].append(field_data)
                                    except:
                                        continue
                                
                                level_data['forms'].append(form_data)
                            except:
                                continue
                        
                        form_levels.append(level_data)
                        
                    except Exception as e:
                        print(f"[WARNING] Failed to analyze level {idx + 1}: {e}")
                        continue
                
                # Navigate back to main page
                try:
                    self.page.goto(original_url, wait_until='domcontentloaded', timeout=15000)
                    self.page.wait_for_load_state('networkidle', timeout=10000)
                except:
                    pass
        
        except Exception as e:
            print(f"[WARNING] Error detecting form levels: {e}")
        
        return form_levels
    
    def _classify_erp_table(self, headers: List[str]) -> str:
        """Classify table type based on headers"""
        headers_text = " ".join([h.lower() for h in headers])
        
        if any(keyword in headers_text for keyword in ['student', 'name', 'roll', 'admission']):
            return 'student_list'
        elif any(keyword in headers_text for keyword in ['fee', 'payment', 'amount', 'balance']):
            return 'fee_record'
        elif any(keyword in headers_text for keyword in ['attendance', 'present', 'absent']):
            return 'attendance'
        elif any(keyword in headers_text for keyword in ['exam', 'test', 'marks', 'grade']):
            return 'examination'
        elif any(keyword in headers_text for keyword in ['teacher', 'staff', 'faculty']):
            return 'teacher_list'
        elif any(keyword in headers_text for keyword in ['book', 'library', 'issue', 'return']):
            return 'library'
        elif any(keyword in headers_text for keyword in ['bus', 'transport', 'route']):
            return 'transport'
        else:
            return 'generic'
    
    def _detect_table_pagination(self, table_locator) -> Dict[str, Any]:
        """Detect pagination controls"""
        pagination = {'exists': False, 'type': None, 'current_page': None, 'total_pages': None}
        
        try:
            # Check for pagination elements
            pagination_selectors = [
                ".pagination",
                "[class*='pagination']",
                "[class*='pager']",
                ".page-info",
                "[class*='page']"
            ]
            
            for selector in pagination_selectors:
                try:
                    pagination_el = table_locator.locator(f".. {selector}, {selector}").first
                    if pagination_el.count():
                        pagination['exists'] = True
                        pagination['type'] = 'standard'
                        break
                except:
                    continue
        except:
            pass
        
        return pagination
    
    def _detect_table_actions(self, table_locator) -> List[str]:
        """Detect action buttons/links in table"""
        actions = []
        
        try:
            action_keywords = ['edit', 'delete', 'view', 'details', 'view', 'action']
            buttons = table_locator.locator("button, a").all()[:10]
            
            for btn in buttons:
                try:
                    text = btn.inner_text().strip().lower()
                    for keyword in action_keywords:
                        if keyword in text and text not in actions:
                            actions.append(text)
                            break
                except:
                    continue
        except:
            pass
        
        return actions
    
    def _has_table_search(self, table_locator) -> bool:
        """Check if table has search functionality"""
        try:
            search_selectors = [
                "input[type='search']",
                "input[placeholder*='search']",
                "input[placeholder*='Search']",
                ".search-input",
                "[class*='search']"
            ]
            
            for selector in search_selectors:
                if table_locator.locator(f".. {selector}, {selector}").first.count():
                    return True
        except:
            pass
        return False
    
    def _has_table_filter(self, table_locator) -> bool:
        """Check if table has filter functionality"""
        try:
            filter_keywords = ['filter', 'Filter', 'FILTER']
            nearby_elements = table_locator.locator("..").first.inner_text().lower()
            return any(keyword.lower() in nearby_elements for keyword in filter_keywords)
        except:
            pass
        return False
    
    def _has_table_export(self, table_locator) -> bool:
        """Check if table has export functionality"""
        try:
            export_keywords = ['export', 'Export', 'download', 'Download', 'excel', 'csv', 'pdf']
            nearby_elements = table_locator.locator("..").first.inner_text().lower()
            return any(keyword.lower() in nearby_elements for keyword in export_keywords)
        except:
            pass
        return False
    
    def find_login_fields(self):
        """Intelligently find login fields"""
        login_keywords = ['username', 'user', 'email', 'login', 'userid', 'user_id', 'usr', 'emailid']
        password_keywords = ['password', 'pass', 'pwd', 'passwd']
        
        username_field = None
        password_field = None
        login_button = None
        
        # Find username field
        for inp in self.detected_elements['input_fields']:
            if inp['type'] in ['text', 'email']:
                name_lower = inp['name'].lower()
                id_lower = (inp['id'] or "").lower()
                placeholder_lower = inp['placeholder'].lower()
                
                if any(keyword in name_lower or keyword in id_lower or keyword in placeholder_lower 
                       for keyword in login_keywords):
                    username_field = inp
                    break
        
        # Find password field
        for inp in self.detected_elements['input_fields']:
            if inp['type'] == 'password':
                password_field = inp
                break
        
        # Find login button
        login_btn_keywords = ['login', 'sign in', 'submit', 'log in', 'signin']
        for btn in self.detected_elements['buttons']:
            btn_text_lower = btn['text'].lower()
            if any(keyword in btn_text_lower for keyword in login_btn_keywords):
                login_button = btn
                break
        
        return username_field, password_field, login_button
    
    def detect_otp_field(self):
        """Detect OTP input field on the page"""
        otp_keywords = ['otp', 'verification', 'verify', 'code', 'pin', '2fa', 'two-factor']
        
        try:
            # Check for input fields with OTP-related attributes
            inputs = self.page.locator("input").all()
            for inp in inputs:
                try:
                    input_id = (inp.get_attribute("id") or "").lower()
                    input_name = (inp.get_attribute("name") or "").lower()
                    input_placeholder = (inp.get_attribute("placeholder") or "").lower()
                    input_type = (inp.get_attribute("type") or "").lower()
                    input_label = ""
                    
                    # Try to find associated label
                    try:
                        label = inp.locator("xpath=ancestor::label | preceding-sibling::label").first
                        if label.count():
                            input_label = label.inner_text().lower()
                    except:
                        pass
                    
                    # Check if any OTP keyword matches
                    combined_text = f"{input_id} {input_name} {input_placeholder} {input_label}"
                    if any(keyword in combined_text for keyword in otp_keywords):
                        return {
                            'id': inp.get_attribute("id"),
                            'name': inp.get_attribute("name"),
                            'type': input_type,
                            'placeholder': inp.get_attribute("placeholder")
                        }
                    
                    # Check for number input types (OTP often uses type="text" or "number")
                    if input_type in ['text', 'number', 'tel']:
                        max_length = inp.get_attribute("maxlength")
                        if max_length and int(max_length) <= 10:  # OTPs are usually short
                            if any(keyword in combined_text for keyword in otp_keywords):
                                return {
                                    'id': inp.get_attribute("id"),
                                    'name': inp.get_attribute("name"),
                                    'type': input_type,
                                    'placeholder': inp.get_attribute("placeholder")
                                }
                except:
                    continue
        except Exception as e:
            print(f"[WARNING] Error detecting OTP field: {e}")
        
        return None
    
    def perform_login(self, otp_value: str = None, wait_for_otp: bool = False):
        """Perform login automatically with optional OTP support"""
        print("[INFO] Attempting to login...")
        
        username_field, password_field, login_button = self.find_login_fields()
        
        if not username_field:
            print("[ERROR] Could not find username field")
            return False
        
        if not password_field:
            print("[ERROR] Could not find password field")
            return False
        
        try:
            # Enter username
            if username_field['id']:
                username_locator = self.page.locator(f"#{username_field['id']}")
            elif username_field['name']:
                username_locator = self.page.locator(f"input[name='{username_field['name']}']")
            else:
                username_locator = self.page.locator(f"input[type='{username_field['type']}']").first
            
            username_locator.fill(self.login_id)
            print(f"[SUCCESS] Entered username: {self.login_id}")
            self.page.wait_for_timeout(500)
            
            # Enter password
            if password_field['id']:
                password_locator = self.page.locator(f"#{password_field['id']}")
            elif password_field['name']:
                password_locator = self.page.locator(f"input[name='{password_field['name']}']")
            else:
                password_locator = self.page.locator("input[type='password']").first
            
            password_locator.fill(self.password)
            print("[SUCCESS] Entered password")
            self.page.wait_for_timeout(500)
            
            # Click login button
            if login_button and login_button['id']:
                login_locator = self.page.locator(f"#{login_button['id']}")
            elif login_button:
                try:
                    login_locator = self.page.locator(f"button:has-text('{login_button['text'][:20]}')").first
                except:
                    login_locator = self.page.locator("button[type='submit'], input[type='submit']").first
            else:
                login_locator = self.page.locator("button[type='submit'], input[type='submit']").first
            
            login_locator.click()
            print("[SUCCESS] Clicked login button")
            self.page.wait_for_load_state('networkidle')
            
            # Check for OTP field after login attempt
            otp_field = self.detect_otp_field()
            if otp_field:
                print("[INFO] OTP field detected! Handling OTP...")
                
                if wait_for_otp:
                    print("[INFO] Waiting for user to enter OTP manually...")
                    # Wait up to 60 seconds for OTP field to be filled
                    try:
                        otp_locator = None
                        if otp_field['id']:
                            otp_locator = self.page.locator(f"#{otp_field['id']}")
                        elif otp_field['name']:
                            otp_locator = self.page.locator(f"input[name='{otp_field['name']}']")
                        else:
                            # Try to find by placeholder
                            if otp_field.get('placeholder'):
                                otp_locator = self.page.locator(f"input[placeholder*='{otp_field['placeholder'][:10]}']").first
                        
                        if otp_locator:
                            # Wait for OTP field to have a value
                            self.page.wait_for_timeout(1000)
                            max_wait = 60  # seconds
                            waited = 0
                            while waited < max_wait:
                                try:
                                    value = otp_locator.input_value()
                                    if value and len(value.strip()) >= 4:  # OTP usually at least 4 digits
                                        print(f"[SUCCESS] OTP entered by user: {value[:2]}**")
                                        break
                                except:
                                    pass
                                time.sleep(1)
                                waited += 1
                            
                            # Click submit/verify button
                            verify_btn = self.page.locator("button:has-text('verify'), button:has-text('submit'), button[type='submit']").first
                            if verify_btn.count():
                                verify_btn.click()
                                self.page.wait_for_load_state('networkidle')
                            else:
                                # Try pressing Enter in OTP field
                                otp_locator.press('Enter')
                                self.page.wait_for_load_state('networkidle')
                    except Exception as e:
                        print(f"[WARNING] Error waiting for OTP: {e}")
                elif otp_value:
                    # Enter provided OTP
                    try:
                        otp_locator = None
                        if otp_field['id']:
                            otp_locator = self.page.locator(f"#{otp_field['id']}")
                        elif otp_field['name']:
                            otp_locator = self.page.locator(f"input[name='{otp_field['name']}']")
                        else:
                            otp_locator = self.page.locator("input[type='text'], input[type='number']").first
                        
                        if otp_locator and otp_locator.count():
                            otp_locator.fill(otp_value)
                            print(f"[SUCCESS] Entered OTP")
                            self.page.wait_for_timeout(500)
                            
                            # Click submit/verify button
                            verify_btn = self.page.locator("button:has-text('verify'), button:has-text('submit'), button[type='submit']").first
                            if verify_btn.count():
                                verify_btn.click()
                                self.page.wait_for_load_state('networkidle')
                            else:
                                otp_locator.press('Enter')
                                self.page.wait_for_load_state('networkidle')
                    except Exception as e:
                        print(f"[ERROR] Error entering OTP: {e}")
                        return False
            
            # Check if login successful
            current_url = self.page.url
            page_source = self.page.content().lower()
            
            if current_url != self.website_url or 'dashboard' in page_source or 'welcome' in page_source or 'home' in page_source:
                print(f"[SUCCESS] Login appears successful! Current URL: {current_url}")
                return True
            else:
                # Check for error messages
                error_indicators = ['error', 'invalid', 'incorrect', 'failed']
                if any(indicator in page_source for indicator in error_indicators):
                    print("[WARNING] Login may have failed - error indicators found")
                    return False
                else:
                    print("[WARNING] Login status unclear")
                    return True
                
        except Exception as e:
            print(f"[ERROR] Login failed: {str(e)}")
            return False
    
    def generate_positive_test_cases(self):
        """Generate positive test cases"""
        print("[INFO] Generating positive test cases...")
        
        # Login positive test - navigate back to original page first
        if self.login_id and self.password:
            if self.page:
                # Navigate back to original URL in case we're on a different page
                try:
                    self.page.goto(self.website_url, wait_until='domcontentloaded')
                    self.page.wait_for_load_state('networkidle')
                except:
                    pass
            login_success = self.perform_login(otp_value=self.otp_value, wait_for_otp=bool(self.otp_value is None and self.page)) if self.page else False
            self.test_cases['positive'].append({
                'test_id': 'POS_LOGIN_001',
                'test_name': 'Valid Login Test',
                'description': 'Verify user can login with valid credentials',
                'steps': [
                    f'Navigate to {self.website_url}',
                    f'Enter username: {self.login_id}',
                    'Enter password: [hidden]',
                    'Click login button',
                    'Verify successful login'
                ],
                'expected_result': 'User should be logged in successfully',
                'status': 'PASS' if login_success else 'FAIL',
                'priority': 'High'
            })
        
        # Form field positive tests
        for inp in self.detected_elements['input_fields']:
            if inp['type'] not in ['password', 'hidden', 'submit', 'button']:
                self.test_cases['positive'].append({
                    'test_id': f"POS_INPUT_{inp['name'].replace(' ', '_')}",
                    'test_name': f"Valid input in {inp['name']} field",
                    'description': f'Verify {inp["name"]} field accepts valid input',
                    'steps': [
                        f"Locate field: {inp['name']}",
                        'Enter valid test data',
                        'Verify input is accepted'
                    ],
                    'expected_result': 'Field should accept valid input',
                    'field': inp['name'],
                    'priority': 'Medium'
                })
        
        # Dropdown positive tests
        for dropdown in self.detected_elements.get('dropdowns', []):
            if not dropdown:
                continue
            label = dropdown.get('name') or dropdown.get('id') or "dropdown"
            sanitized = self._sanitize_identifier(label)
            self.test_cases['positive'].append({
                'test_id': f"POS_DROPDOWN_{sanitized}",
                'test_name': f"Valid selection in {label} dropdown",
                'description': f'Ensure dropdown {label} accepts valid selections',
                'steps': [
                    f"Locate dropdown: {label}",
                    'Open dropdown options',
                    'Select a valid option',
                    'Verify selection is applied'
                ],
                'expected_result': 'Dropdown should update value with selected option',
                'field': label,
                'priority': 'Medium'
            })
        
        print(f"[SUCCESS] Generated {len(self.test_cases['positive'])} positive test cases")
    
    def generate_negative_test_cases(self):
        """Generate negative test cases"""
        print("[INFO] Generating negative test cases...")
        
        # Login negative tests
        negative_scenarios = [
            {
                'test_id': 'NEG_LOGIN_001',
                'test_name': 'Login with empty username',
                'description': 'Verify error message when username is empty',
                'steps': [
                    'Navigate to login page',
                    'Leave username field empty',
                    'Enter password',
                    'Click login button'
                ],
                'expected_result': 'Error message should be displayed',
                'priority': 'High'
            },
            {
                'test_id': 'NEG_LOGIN_002',
                'test_name': 'Login with empty password',
                'description': 'Verify error message when password is empty',
                'steps': [
                    'Navigate to login page',
                    'Enter username',
                    'Leave password field empty',
                    'Click login button'
                ],
                'expected_result': 'Error message should be displayed',
                'priority': 'High'
            },
            {
                'test_id': 'NEG_LOGIN_003',
                'test_name': 'Login with invalid credentials',
                'description': 'Verify error message for invalid credentials',
                'steps': [
                    'Navigate to login page',
                    'Enter invalid username: invalid_user',
                    'Enter invalid password: invalid_pass',
                    'Click login button'
                ],
                'expected_result': 'Invalid credentials error should be displayed',
                'priority': 'High'
            },
            {
                'test_id': 'NEG_LOGIN_004',
                'test_name': 'Login with SQL injection attempt',
                'description': 'Verify system handles SQL injection attempts',
                'steps': [
                    'Navigate to login page',
                    'Enter username: admin\' OR \'1\'=\'1',
                    'Enter password: test',
                    'Click login button'
                ],
                'expected_result': 'SQL injection should be blocked',
                'priority': 'Critical'
            },
            {
                'test_id': 'NEG_LOGIN_005',
                'test_name': 'Login with XSS attempt',
                'description': 'Verify system handles XSS attacks',
                'steps': [
                    'Navigate to login page',
                    'Enter username: <script>alert("XSS")</script>',
                    'Enter password: test',
                    'Click login button'
                ],
                'expected_result': 'XSS attack should be prevented',
                'priority': 'Critical'
            }
        ]
        
        self.test_cases['negative'].extend(negative_scenarios)
        
        # Input field negative tests
        for inp in self.detected_elements['input_fields']:
            if inp['required']:
                self.test_cases['negative'].append({
                    'test_id': f"NEG_INPUT_{inp['name'].replace(' ', '_')}_EMPTY",
                    'test_name': f"Empty {inp['name']} field test",
                    'description': f'Verify {inp["name"]} field validation for empty input',
                    'steps': [
                        f"Leave {inp['name']} field empty",
                        'Submit form',
                        'Verify validation error'
                    ],
                    'expected_result': 'Validation error should be displayed',
                    'priority': 'Medium'
                })
            
            # Boundary value tests
            if inp['type'] in ['text', 'number', 'email']:
                self.test_cases['negative'].append({
                    'test_id': f"NEG_INPUT_{inp['name'].replace(' ', '_')}_BOUNDARY",
                    'test_name': f"Boundary value test for {inp['name']}",
                    'description': f'Test {inp["name"]} with boundary values',
                    'steps': [
                        f"Enter very long text in {inp['name']} field",
                        'Submit form',
                        'Verify validation'
                    ],
                    'expected_result': 'Field should handle boundary values correctly',
                    'priority': 'Medium'
                })
        
        # Security payload tests for text inputs
        security_inputs = [
            inp for inp in self.detected_elements['input_fields']
            if inp['type'] in ['text', 'email', 'search', 'textarea', 'url']
        ]
        payloads = [
            ('SQLI', "SQL injection attempt", "admin' OR '1'='1"),
            ('XSS', "XSS script injection", '<script>alert(\"XSS\")</script>')
        ]
        for inp in security_inputs[:10]:
            label = inp['name'] or inp['id'] or "field"
            sanitized = self._sanitize_identifier(label)
            for suffix, description, payload in payloads:
                self.test_cases['negative'].append({
                    'test_id': f"NEG_SECURITY_{sanitized}_{suffix}",
                    'test_name': f"{description} in {label}",
                    'description': f'Test how {label} handles {description.lower()} payloads',
                    'steps': [
                        f"Locate field: {label}",
                        f"Enter payload: {payload}",
                        'Submit the enclosing form or action',
                        'Observe application response'
                    ],
                    'expected_result': 'Application should reject or sanitize the malicious payload',
                    'priority': 'Critical',
                    'field': inp['name'],
                    'field_id': inp['id'],
                    'malicious_payload': payload
                })
        
        print(f"[SUCCESS] Generated {len(self.test_cases['negative'])} negative test cases")
    
    def generate_ui_test_cases(self):
        """Generate UI test cases"""
        print("[INFO] Generating UI test cases...")
        
        # Button UI tests
        for btn in self.detected_elements['buttons']:
            self.test_cases['ui'].append({
                'test_id': f"UI_BUTTON_{btn['text'].replace(' ', '_').replace('/', '_')[:30]}",
                'test_name': f"Verify {btn['text']} button is visible",
                'description': f'Check if {btn["text"]} button is displayed correctly',
                'steps': [
                    f"Locate button: {btn['text']}",
                    'Verify button is visible',
                    'Verify button is clickable',
                    'Verify button styling is consistent'
                ],
                'expected_result': 'Button should be visible, clickable, and properly styled',
                'priority': 'Medium'
            })
        
        # Input field UI tests
        for inp in self.detected_elements['input_fields']:
            if inp['type'] not in ['hidden', 'submit']:
                self.test_cases['ui'].append({
                    'test_id': f"UI_INPUT_{inp['name'].replace(' ', '_')}",
                    'test_name': f"Verify {inp['name']} field UI",
                    'description': f'Check UI of {inp["name"]} input field',
                    'steps': [
                        f"Locate field: {inp['name']}",
                        'Verify field is visible',
                        'Verify placeholder text is displayed (if applicable)',
                        'Verify field styling is consistent',
                        'Verify field is properly aligned'
                    ],
                    'expected_result': 'Field should be properly styled, visible, and aligned',
                    'priority': 'Medium'
                })
        
        # Checkbox UI tests
        for checkbox in self.detected_elements.get('checkboxes', []):
            label = checkbox.get('label') or checkbox.get('name') or "Checkbox"
            sanitized = self._sanitize_identifier(label)
            self.test_cases['ui'].append({
                'test_id': f"UI_CHECKBOX_{sanitized}",
                'test_name': f"Verify {label} checkbox visibility",
                'description': f'Ensure {label} checkbox is visible, aligned, and toggleable',
                'steps': [
                    f"Locate checkbox: {label}",
                    'Verify checkbox is visible',
                    'Toggle checkbox on and off',
                    'Verify state changes correctly'
                ],
                'expected_result': 'Checkbox should be visible and reflect user interaction',
                'priority': 'Medium'
            })
        
        # Radio button UI tests
        for radio in self.detected_elements.get('radio_buttons', []):
            label = radio.get('label') or radio.get('name') or "Radio option"
            sanitized = self._sanitize_identifier(label)
            self.test_cases['ui'].append({
                'test_id': f"UI_RADIO_{sanitized}",
                'test_name': f"Verify {label} radio option",
                'description': f'Ensure radio option {label} is visible and selectable',
                'steps': [
                    f"Locate radio option: {label}",
                    'Verify option is visible',
                    'Select the option and verify other options respond appropriately'
                ],
                'expected_result': 'Radio option should be selectable and mutually exclusive',
                'priority': 'Medium'
            })
        
        # Iframe UI tests
        for frame in self.detected_elements.get('iframes', []):
            frame_label = frame.get('id') or frame.get('name') or frame.get('src') or "iframe"
            sanitized = self._sanitize_identifier(frame_label)
            self.test_cases['ui'].append({
                'test_id': f"UI_IFRAME_{sanitized}",
                'test_name': f"Verify iframe {frame_label} visibility",
                'description': f'Ensure iframe {frame_label} loads correctly',
                'steps': [
                    f"Locate iframe: {frame_label}",
                    'Verify iframe is visible',
                    'Verify iframe content loads without errors'
                ],
                'expected_result': 'Iframe content should be accessible and visible',
                'priority': 'Low'
            })
        
        # Page UI tests
        self.test_cases['ui'].append({
            'test_id': 'UI_PAGE_001',
            'test_name': 'Verify page title',
            'description': 'Check if page title is displayed correctly',
            'steps': [
                'Load the page',
                'Verify page title is present',
                'Verify page title is meaningful'
            ],
            'expected_result': f'Page title should be: {self.detected_elements.get("page_title", "N/A")}',
            'priority': 'Low'
        })
        
        self.test_cases['ui'].append({
            'test_id': 'UI_PAGE_002',
            'test_name': 'Verify responsive design',
            'description': 'Check if page is responsive on different screen sizes',
            'steps': [
                'Open page on desktop',
                'Resize browser to mobile size',
                'Verify layout adjusts properly',
                'Verify all elements are accessible'
            ],
            'expected_result': 'Page should be responsive and usable on all screen sizes',
            'priority': 'High'
        })
        
        print(f"[SUCCESS] Generated {len(self.test_cases['ui'])} UI test cases")
    
    def generate_functional_test_cases(self):
        """Generate functional test cases"""
        print("[INFO] Generating functional test cases...")
        
        # Navigation tests
        for link in self.detected_elements['links'][:15]:
            self.test_cases['functional'].append({
                'test_id': f"FUNC_NAV_{link['text'].replace(' ', '_').replace('/', '_')[:25]}",
                'test_name': f"Verify navigation to {link['text']}",
                'description': f'Test navigation functionality for {link["text"]} link',
                'steps': [
                    f"Navigate to page: {self.website_url}",
                    f"Locate and click on link: {link['text']}",
                    'Wait for page to load',
                    'Verify page navigation occurs',
                    'Verify new page URL is correct',
                    'Verify page loads without errors',
                    'Verify page content is displayed',
                    'Verify no broken links or missing resources'
                ],
                'expected_result': 'Page should navigate successfully without errors and display correct content',
                'link': link['href'],
                'priority': 'Medium',
                'test_type': 'navigation'
            })
        
        # Button functionality tests
        for btn in self.detected_elements['buttons']:
            if btn['text'].lower() not in ['login', 'submit']:  # Skip login as it's already tested
                self.test_cases['functional'].append({
                    'test_id': f"FUNC_BTN_{btn['text'].replace(' ', '_').replace('/', '_')[:25]}",
                    'test_name': f"Verify {btn['text']} button functionality",
                    'description': f'Test functionality of {btn["text"]} button',
                    'steps': [
                        f"Navigate to page: {self.website_url}",
                        f"Locate button: {btn['text']}",
                        'Verify button is visible and enabled',
                        f"Click button: {btn['text']}",
                        'Wait for action to complete',
                        'Verify expected action occurs',
                        'Verify no error messages are displayed',
                        'Verify page state changes if applicable'
                    ],
                    'expected_result': 'Button should perform expected action without errors',
                    'priority': 'Medium',
                    'test_type': 'button_functionality'
                })
        
        # Form submission tests
        for form in self.detected_elements['forms']:
            self.test_cases['functional'].append({
                'test_id': f"FUNC_FORM_{form.get('id', 'unknown')}",
                'test_name': f"Verify form submission functionality",
                'description': f'Test form submission for form: {form.get("id", "Unknown")}',
                'steps': [
                    f"Navigate to page: {self.website_url}",
                    'Locate the form',
                    'Fill all required fields with valid data',
                    'Verify all fields are filled correctly',
                    'Submit the form',
                    'Wait for form submission to complete',
                    'Verify form submission is successful',
                    'Verify appropriate response is received',
                    'Verify success message or redirect occurs'
                ],
                'expected_result': 'Form should submit successfully with proper validation',
                'priority': 'High',
                'test_type': 'form_submission'
            })
        
        # Input field functionality tests
        for inp in self.detected_elements['input_fields']:
            if inp['type'] not in ['hidden', 'submit', 'button', 'password']:
                self.test_cases['functional'].append({
                    'test_id': f"FUNC_INPUT_{inp['name'].replace(' ', '_')}",
                    'test_name': f"Verify {inp['name']} input field functionality",
                    'description': f'Test input functionality for {inp["name"]} field',
                    'steps': [
                        f"Navigate to page: {self.website_url}",
                        f"Locate input field: {inp['name']}",
                        'Verify field is visible and enabled',
                        'Enter test data in the field',
                        'Verify data is accepted',
                        'Verify field validation works',
                        'Verify field formatting if applicable'
                    ],
                    'expected_result': 'Input field should accept and process data correctly',
                    'priority': 'Medium',
                    'test_type': 'input_functionality'
                })
        
        # Dropdown functionality tests
        for dd in self.detected_elements['dropdowns']:
            self.test_cases['functional'].append({
                'test_id': f"FUNC_DROPDOWN_{dd.get('name', 'unknown').replace(' ', '_')}",
                'test_name': f"Verify dropdown {dd.get('name', 'Unknown')} functionality",
                'description': f'Test dropdown functionality for {dd.get("name", "Unknown")}',
                'steps': [
                    f"Navigate to page: {self.website_url}",
                    f"Locate dropdown: {dd.get('name', 'Unknown')}",
                    'Click on dropdown to open',
                    'Verify dropdown options are displayed',
                    'Select an option from dropdown',
                    'Verify selected option is displayed',
                    'Verify dropdown value is updated'
                ],
                'expected_result': 'Dropdown should allow selection and update value correctly',
                'priority': 'Medium',
                'test_type': 'dropdown_functionality'
            })
        
        # Table validation tests
        for table in self.detected_elements.get('tables', []):
            table_id = table.get('id') or "data_table"
            self.test_cases['functional'].append({
                'test_id': f"FUNC_TABLE_{self._sanitize_identifier(table_id)}",
                'test_name': f"Verify data table {table_id} renders correctly",
                'description': f'Check table {table_id} row counts and headers',
                'steps': [
                    f"Locate table: {table_id}",
                    'Verify headers are displayed',
                    'Verify row count matches expected data',
                    'Scroll through table content if paginated'
                ],
                'expected_result': 'Table should render rows and headers without layout issues',
                'priority': 'Medium',
                'test_type': 'table_validation',
                'table': table
            })
        
        # Iframe interaction tests
        for frame in self.detected_elements.get('iframes', []):
            frame_label = frame.get('id') or frame.get('name') or frame.get('src') or "iframe"
            self.test_cases['functional'].append({
                'test_id': f"FUNC_IFRAME_{self._sanitize_identifier(frame_label)}",
                'test_name': f"Verify iframe {frame_label} interaction",
                'description': f'Ensure iframe {frame_label} loads and can be switched to',
                'steps': [
                    f"Locate iframe: {frame_label}",
                    'Switch focus to iframe',
                    'Verify iframe content loads',
                    'Switch back to parent context'
                ],
                'expected_result': 'Iframe should be accessible and load content.',
                'priority': 'Low',
                'test_type': 'iframe_validation',
                'iframe': frame
            })
        
        # Page load and performance tests
        self.test_cases['functional'].append({
            'test_id': 'FUNC_PERF_001',
            'test_name': 'Verify page load performance',
            'description': 'Test page load time and performance',
            'steps': [
                f"Navigate to page: {self.website_url}",
                'Measure page load time',
                'Verify page loads within acceptable time',
                'Verify all resources are loaded',
                'Verify no broken images or missing resources'
            ],
            'expected_result': 'Page should load within acceptable time with all resources',
            'priority': 'Medium',
            'test_type': 'performance'
        })
        
        # Generate multi-step workflows
        self.generate_multi_step_workflows()
        
        # Generate edge cases
        self.generate_edge_case_tests()
        
        print(f"[SUCCESS] Generated {len(self.test_cases['functional'])} functional test cases")
    
    def generate_multi_step_workflows(self):
        """Generate context-aware multi-step workflow test cases"""
        print("[INFO] Generating multi-step workflows...")
        
        # Login workflow (if credentials provided)
        if self.login_id and self.password:
            self.test_cases['workflow'].append({
                'test_id': 'WORKFLOW_LOGIN_001',
                'test_name': 'Complete Login Workflow',
                'description': 'Test end-to-end login workflow with all steps',
                'steps': [
                    f"Navigate to login page: {self.website_url}",
                    'Verify login page is displayed',
                    'Verify username field is visible',
                    'Verify password field is visible',
                    'Enter valid username',
                    'Enter valid password',
                    'Click login button',
                    'Wait for authentication',
                    'Handle OTP if required',
                    'Verify successful login',
                    'Verify user is redirected to appropriate page',
                    'Verify user session is established',
                    'Verify dashboard/widgets are loaded',
                    'Verify user can access protected resources'
                ],
                'expected_result': 'Complete login workflow should work seamlessly',
                'priority': 'Critical',
                'test_type': 'workflow',
                'category': 'authentication',
                'estimated_time': '2-3 minutes'
            })
        
        # Form submission workflow
        forms = self.detected_elements.get('forms', [])
        if forms:
            for idx, form in enumerate(forms[:3]):  # Limit to first 3 forms
                self.test_cases['workflow'].append({
                    'test_id': f'WORKFLOW_FORM_{idx + 1:03d}',
                    'test_name': f'Complete Form Submission Workflow - Form {idx + 1}',
                    'description': f'Test end-to-end form submission workflow for form {idx + 1}',
                    'steps': [
                        f"Navigate to page: {self.website_url}",
                        'Locate form on page',
                        'Verify all required fields are visible',
                        'Fill all mandatory fields with valid data',
                        'Fill optional fields if applicable',
                        'Verify form validation (if any)',
                        'Submit form',
                        'Wait for submission to complete',
                        'Verify success message/confirmation',
                        'Verify redirect or page update',
                        'Verify submitted data appears correctly'
                    ],
                    'expected_result': 'Form should submit successfully and display confirmation',
                    'priority': 'High',
                    'test_type': 'workflow',
                    'category': 'form_submission',
                    'estimated_time': '1-2 minutes'
                })
        
        # Module navigation workflow for School ERP
        modules = self.detected_elements.get('modules', [])
        for module in modules[:3]:  # Limit to first 3 modules
            module_name = module.get('name', '').lower()
            module_links = module.get('links', [])[:3]  # First 3 links per module
            
            if module_links:
                self.test_cases['workflow'].append({
                    'test_id': f'WORKFLOW_MODULE_{self._sanitize_identifier(module_name)}',
                    'test_name': f'Module Navigation Workflow - {module_name.title()}',
                    'description': f'Test navigation through {module_name} module',
                    'steps': [
                        'Login to system (if required)',
                        f'Navigate to {module_name} module',
                        'Verify module page loads correctly',
                        f'Click on "{module_links[0].get("text", "First Link")}" link',
                        'Verify navigation occurs',
                        'Verify page content is correct',
                        'Navigate back to module main page',
                        'Verify all module links are accessible'
                    ],
                    'expected_result': f'All navigation within {module_name} module should work correctly',
                    'priority': 'High',
                    'test_type': 'workflow',
                    'category': 'module_navigation',
                    'module': module_name,
                    'estimated_time': '2-3 minutes'
                })
    
    def generate_edge_case_tests(self):
        """Generate edge case test scenarios"""
        print("[INFO] Generating edge case tests...")
        
        # Boundary value tests for input fields
        input_fields = self.detected_elements.get('input_fields', [])
        for inp in input_fields[:10]:  # Limit to first 10 fields
            if inp['type'] in ['text', 'number', 'email']:
                field_name = inp.get('name', 'field')
                
                # Maximum length test
                self.test_cases['edge_case'].append({
                    'test_id': f'EDGE_MAX_LENGTH_{self._sanitize_identifier(field_name)}',
                    'test_name': f'Maximum Length Test - {field_name}',
                    'description': f'Test field accepts maximum allowed characters',
                    'steps': [
                        f"Navigate to page: {self.website_url}",
                        f'Locate field: {field_name}',
                        'Generate input with maximum allowed length',
                        'Enter maximum length value',
                        'Verify field accepts the input',
                        'Verify field truncates or validates correctly'
                    ],
                    'expected_result': 'Field should handle maximum length input correctly',
                    'priority': 'Medium',
                    'test_type': 'boundary',
                    'category': 'input_validation'
                })
                
                # Empty string test
                if inp.get('required'):
                    self.test_cases['edge_case'].append({
                        'test_id': f'EDGE_EMPTY_REQUIRED_{self._sanitize_identifier(field_name)}',
                        'test_name': f'Empty Required Field Test - {field_name}',
                        'description': f'Test validation when required field is empty',
                        'steps': [
                            f"Navigate to page: {self.website_url}",
                            f'Locate required field: {field_name}',
                            'Leave field empty',
                            'Attempt to submit form',
                            'Verify error message is displayed',
                            'Verify form does not submit'
                        ],
                        'expected_result': 'Required field should show validation error when empty',
                        'priority': 'High',
                        'test_type': 'validation',
                        'category': 'required_field'
                    })
                
                # Special characters test
                self.test_cases['edge_case'].append({
                    'test_id': f'EDGE_SPECIAL_CHARS_{self._sanitize_identifier(field_name)}',
                    'test_name': f'Special Characters Test - {field_name}',
                    'description': f'Test field handling of special characters',
                    'steps': [
                        f"Navigate to page: {self.website_url}",
                        f'Locate field: {field_name}',
                        'Enter value with special characters (!@#$%^&*)',
                        'Verify field accepts or rejects correctly',
                        'Verify proper sanitization or validation'
                    ],
                    'expected_result': 'Field should handle special characters appropriately',
                    'priority': 'Medium',
                    'test_type': 'validation',
                    'category': 'special_characters'
                })
        
        # Table edge cases
        tables = self.detected_elements.get('table_details', [])
        for table in tables[:5]:  # Limit to first 5 tables
            if table.get('pagination', {}).get('exists'):
                self.test_cases['edge_case'].append({
                    'test_id': f'EDGE_TABLE_PAGINATION_{table.get("id", "table")}',
                    'test_name': f'Table Pagination Edge Case - {table.get("id", "Table")}',
                    'description': f'Test table pagination edge cases (first, last, empty)',
                    'steps': [
                        f"Navigate to page: {self.website_url}",
                        f'Locate table: {table.get("id")}',
                        'Navigate to first page',
                        'Verify first page displays correctly',
                        'Navigate to last page',
                        'Verify last page displays correctly',
                        'Test empty table scenario (if applicable)',
                        'Verify pagination controls work correctly'
                    ],
                    'expected_result': 'Table pagination should handle edge cases correctly',
                    'priority': 'Medium',
                    'test_type': 'pagination',
                    'category': 'table_behavior'
                })
        
        # Concurrent action tests
        if self.detected_elements.get('buttons'):
            self.test_cases['edge_case'].append({
                'test_id': 'EDGE_RAPID_CLICKS',
                'test_name': 'Rapid Button Clicks Test',
                'description': 'Test behavior when button is clicked rapidly multiple times',
                'steps': [
                    f"Navigate to page: {self.website_url}",
                    'Locate a submit/submit-like button',
                    'Click button rapidly 3-5 times',
                    'Verify system handles rapid clicks correctly',
                    'Verify no duplicate submissions occur',
                    'Verify button state updates correctly'
                ],
                'expected_result': 'System should prevent duplicate submissions from rapid clicks',
                'priority': 'Medium',
                'test_type': 'concurrency',
                'category': 'button_behavior'
            })
        
        # Session timeout test (if login exists)
        if self.login_id and self.password:
            self.test_cases['edge_case'].append({
                'test_id': 'EDGE_SESSION_TIMEOUT',
                'test_name': 'Session Timeout Test',
                'description': 'Test behavior when session expires',
                'steps': [
                    'Login to system',
                    'Wait for session to timeout (or simulate timeout)',
                    'Attempt to perform an action',
                    'Verify session expired message is displayed',
                    'Verify user is redirected to login page',
                    'Verify user can login again after timeout'
                ],
                'expected_result': 'System should handle session timeout gracefully',
                'priority': 'High',
                'test_type': 'session',
                'category': 'authentication'
            })
    
    def _sanitize_identifier(self, text: str) -> str:
        if not text:
            return "FIELD"
        value = re.sub(r'[^a-zA-Z0-9]+', '_', text).strip('_')
        return value or "FIELD"
    
    def generate_module_specific_tests(self):
        """Generate test cases for detected School ERP modules"""
        if not self.detected_elements.get('modules'):
            return
        
        print("[INFO] Generating module-specific test cases...")
        
        modules = self.detected_elements.get('modules', [])
        for module in modules:
            module_name = module.get('name', '').lower()
            
            if module_name == 'student':
                self._generate_student_module_tests(module)
            elif module_name == 'attendance':
                self._generate_attendance_module_tests(module)
            elif module_name == 'examination':
                self._generate_examination_module_tests(module)
            elif module_name == 'finance':
                self._generate_finance_module_tests(module)
            elif module_name == 'teacher':
                self._generate_teacher_module_tests(module)
    
    def _generate_student_module_tests(self, module: Dict):
        """Generate tests for Student Management module"""
        module_links = module.get('links', [])
        for link in module_links[:5]:  # Limit to first 5 links
            link_text = link.get('text', '')
            self.test_cases['functional'].append({
                'test_id': f"FUNC_STUDENT_{self._sanitize_identifier(link_text)}",
                'test_name': f"Student Module: Navigate to {link_text}",
                'description': f'Verify navigation to {link_text} in Student module',
                'test_type': 'navigation',
                'link': link.get('href', ''),
                'module': 'student',
                'priority': 'High'
            })
    
    def _generate_attendance_module_tests(self, module: Dict):
        """Generate tests for Attendance module"""
        module_links = module.get('links', [])
        for link in module_links[:5]:
            link_text = link.get('text', '')
            self.test_cases['functional'].append({
                'test_id': f"FUNC_ATTENDANCE_{self._sanitize_identifier(link_text)}",
                'test_name': f"Attendance Module: Navigate to {link_text}",
                'description': f'Verify navigation to {link_text} in Attendance module',
                'test_type': 'navigation',
                'link': link.get('href', ''),
                'module': 'attendance',
                'priority': 'High'
            })
        
        # Add table tests for attendance records
        tables = self.detected_elements.get('table_details', [])
        for table in tables:
            if table.get('type') == 'attendance':
                self.test_cases['functional'].append({
                    'test_id': 'FUNC_ATTENDANCE_TABLE_VIEW',
                    'test_name': 'Attendance Table: Verify table display',
                    'description': 'Verify attendance table displays correctly',
                    'test_type': 'table_validation',
                    'table': table,
                    'module': 'attendance',
                    'priority': 'Medium'
                })
    
    def _generate_examination_module_tests(self, module: Dict):
        """Generate tests for Examination module"""
        module_links = module.get('links', [])
        for link in module_links[:5]:
            link_text = link.get('text', '')
            self.test_cases['functional'].append({
                'test_id': f"FUNC_EXAM_{self._sanitize_identifier(link_text)}",
                'test_name': f"Examination Module: Navigate to {link_text}",
                'description': f'Verify navigation to {link_text} in Examination module',
                'test_type': 'navigation',
                'link': link.get('href', ''),
                'module': 'examination',
                'priority': 'High'
            })
        
        # Add tests for exam result tables
        tables = self.detected_elements.get('table_details', [])
        for table in tables:
            if table.get('type') == 'examination':
                self.test_cases['functional'].append({
                    'test_id': 'FUNC_EXAM_TABLE_VIEW',
                    'test_name': 'Examination Table: Verify results display',
                    'description': 'Verify examination results table displays correctly',
                    'test_type': 'table_validation',
                    'table': table,
                    'module': 'examination',
                    'priority': 'High'
                })
    
    def _generate_finance_module_tests(self, module: Dict):
        """Generate tests for Finance/Fee module"""
        module_links = module.get('links', [])
        for link in module_links[:5]:
            link_text = link.get('text', '')
            self.test_cases['functional'].append({
                'test_id': f"FUNC_FINANCE_{self._sanitize_identifier(link_text)}",
                'test_name': f"Finance Module: Navigate to {link_text}",
                'description': f'Verify navigation to {link_text} in Finance module',
                'test_type': 'navigation',
                'link': link.get('href', ''),
                'module': 'finance',
                'priority': 'High'
            })
        
        # Add tests for fee record tables
        tables = self.detected_elements.get('table_details', [])
        for table in tables:
            if table.get('type') == 'fee_record':
                self.test_cases['functional'].append({
                    'test_id': 'FUNC_FEE_TABLE_VIEW',
                    'test_name': 'Fee Table: Verify fee records display',
                    'description': 'Verify fee records table displays correctly',
                    'test_type': 'table_validation',
                    'table': table,
                    'module': 'finance',
                    'priority': 'High'
                })
    
    def _generate_teacher_module_tests(self, module: Dict):
        """Generate tests for Teacher module"""
        module_links = module.get('links', [])
        for link in module_links[:5]:
            link_text = link.get('text', '')
            self.test_cases['functional'].append({
                'test_id': f"FUNC_TEACHER_{self._sanitize_identifier(link_text)}",
                'test_name': f"Teacher Module: Navigate to {link_text}",
                'description': f'Verify navigation to {link_text} in Teacher module',
                'test_type': 'navigation',
                'link': link.get('href', ''),
                'module': 'teacher',
                'priority': 'Medium'
            })
    
    def _sanitize_identifier(self, text: str) -> str:
        """Convert text to a valid identifier"""
        if not text:
            return "UNKNOWN"
        # Remove special characters and spaces
        sanitized = re.sub(r'[^a-zA-Z0-9]', '_', text.strip())
        # Remove multiple underscores
        sanitized = re.sub(r'_+', '_', sanitized)
        # Remove leading/trailing underscores
        sanitized = sanitized.strip('_')
        # Convert to uppercase and limit length
        return sanitized.upper()[:30] if sanitized else "UNKNOWN"
    
    def generate_all_tests(self):
        """Generate all types of test cases"""
        if not self.page:
            self.initialize_driver()
        
        self.analyze_website()
        
        # Generate all test types
        self.generate_positive_test_cases()
        self.generate_negative_test_cases()
        self.generate_ui_test_cases()
        self.generate_functional_test_cases()
        
        # Generate module-specific tests for School ERP
        self.generate_module_specific_tests()
        
        # Generate multi-level form tests
        if self.detected_elements.get('form_levels'):
            self.test_cases['multi_level'] = self._generate_multi_level_tests()
        
        # Note: Multi-step workflows and edge cases are generated within generate_functional_test_cases()
        # They are stored in self.test_cases['workflow'] and self.test_cases['edge_case']
        
        return self.test_cases
    
    def _generate_multi_level_tests(self) -> List[Dict[str, Any]]:
        """Generate test cases for multi-level forms"""
        multi_level_tests = []
        form_levels = self.detected_elements.get('form_levels', [])
        
        if not form_levels:
            return multi_level_tests
        
        print(f"[INFO] Generating test cases for {len(form_levels)} form levels...")
        
        for level_data in form_levels:
            level_num = level_data.get('level_number', 0)
            level_name = level_data.get('level_name', f'Level {level_num}')
            level_url = level_data.get('url', '')
            forms = level_data.get('forms', [])
            
            if not forms:
                continue
            
            # Generate positive tests for each form in this level
            for form_idx, form in enumerate(forms):
                form_id = form.get('form_id', f'form_{form_idx + 1}')
                input_fields = form.get('input_fields', [])
                
                # Positive test - form submission with valid data
                test_id = f"MULTI_LEVEL_{level_num}_FORM_{form_idx + 1:03d}_POSITIVE"
                
                multi_level_tests.append({
                    'test_id': test_id,
                    'test_name': f'Test {level_name} - Form {form_idx + 1} (Positive)',
                    'description': f'Test form submission for {level_name} with valid data',
                    'level_number': level_num,
                    'level_name': level_name,
                    'level_url': level_url,
                    'form_id': form_id,
                    'form_action': form.get('action', ''),
                    'form_method': form.get('method', 'get'),
                    'steps': [
                        f"Navigate to {level_url}",
                        "Wait for page to load completely",
                        f"Locate form: {form_id}",
                        "Verify form is visible and accessible",
                        "Fill all required fields with valid test data",
                        "Fill optional fields if present",
                        "Submit form",
                        "Wait for submission to complete",
                        "Verify success message or confirmation"
                    ],
                    'expected_result': f'Form in {level_name} should submit successfully with valid data',
                    'priority': 'High',
                    'test_type': 'multi_level_form',
                    'category': 'form_submission',
                    'field_count': len(input_fields),
                    'required_field_count': len([f for f in input_fields if f.get('required', False)])
                })
                
                # Negative test - form submission with invalid data
                test_id_neg = f"MULTI_LEVEL_{level_num}_FORM_{form_idx + 1:03d}_NEGATIVE"
                
                multi_level_tests.append({
                    'test_id': test_id_neg,
                    'test_name': f'Test {level_name} - Form {form_idx + 1} (Negative)',
                    'description': f'Test form validation for {level_name} with invalid data',
                    'level_number': level_num,
                    'level_name': level_name,
                    'level_url': level_url,
                    'form_id': form_id,
                    'form_action': form.get('action', ''),
                    'form_method': form.get('method', 'get'),
                    'steps': [
                        f"Navigate to {level_url}",
                        "Wait for page to load completely",
                        f"Locate form: {form_id}",
                        "Leave required fields empty",
                        "Submit form",
                        "Verify validation error messages appear",
                        "Verify form does not submit"
                    ],
                    'expected_result': f'Form in {level_name} should show validation errors for invalid data',
                    'priority': 'Medium',
                    'test_type': 'multi_level_form',
                    'category': 'form_validation',
                    'field_count': len(input_fields),
                    'required_field_count': len([f for f in input_fields if f.get('required', False)])
                })
                
                # UI test - verify all form fields are present
                test_id_ui = f"MULTI_LEVEL_{level_num}_FORM_{form_idx + 1:03d}_UI"
                
                multi_level_tests.append({
                    'test_id': test_id_ui,
                    'test_name': f'Test {level_name} - Form {form_idx + 1} (UI Verification)',
                    'description': f'Verify UI elements and layout of form in {level_name}',
                    'level_number': level_num,
                    'level_name': level_name,
                    'level_url': level_url,
                    'form_id': form_id,
                    'steps': [
                        f"Navigate to {level_url}",
                        "Wait for page to load completely",
                        f"Locate form: {form_id}",
                        "Verify form is visible",
                        f"Verify all {len(input_fields)} form fields are present",
                        "Verify labels are displayed correctly",
                        "Verify submit button is visible and enabled"
                    ],
                    'expected_result': f'All form elements in {level_name} should be visible and properly displayed',
                    'priority': 'Medium',
                    'test_type': 'multi_level_form',
                    'category': 'ui_verification',
                    'field_count': len(input_fields)
                })
        
        print(f"[INFO] Generated {len(multi_level_tests)} multi-level test cases")
        return multi_level_tests
    
    def save_report(self):
        """Save test report to file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = f"test_report_{timestamp}.json"
        report_path = os.path.join('reports', report_file)
        
        report_data = {
            'website_url': self.website_url,
            'generated_at': datetime.now().isoformat(),
            'test_cases': self.test_cases,
            'summary': {
                'total_tests': sum(len(tests) for tests in self.test_cases.values()),
                'positive': len(self.test_cases.get('positive', [])),
                'negative': len(self.test_cases.get('negative', [])),
                'ui': len(self.test_cases.get('ui', [])),
                'functional': len(self.test_cases.get('functional', [])),
                'workflow': len(self.test_cases.get('workflow', [])),
                'edge_case': len(self.test_cases.get('edge_case', [])),
                'multi_level': len(self.test_cases.get('multi_level', []))
            },
            'detected_elements': {
                'input_fields': len(self.detected_elements.get('input_fields', [])),
                'buttons': len(self.detected_elements.get('buttons', [])),
                'links': len(self.detected_elements.get('links', [])),
                'forms': len(self.detected_elements.get('forms', []))
            }
        }
        
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False, default=str)
        
        return report_file
    
    def save_test_cases_excel(self) -> str:
        """Export all generated test cases to Excel with English descriptions"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_file = f"test_cases_{timestamp}.xlsx"
        xlsx_path = os.path.join('reports', xlsx_file)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            # Fallback to CSV if openpyxl isn't installed
            import csv
            csv_file = f"test_cases_{timestamp}.csv"
            csv_path = os.path.join('reports', csv_file)
            os.makedirs('reports', exist_ok=True)
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'Test ID', 'Test Category', 'Test Name', 'Description (English)', 
                    'Test Steps', 'Expected Result', 'Priority', 'Field/Element', 'Status', 'Result', 'Error Message'
                ])
                
                # Write all test cases
                for category, test_list in self.test_cases.items():
                    for test in test_list:
                        steps_str = ' | '.join(test.get('steps', [])) if isinstance(test.get('steps'), list) else str(test.get('steps', ''))
                        writer.writerow([
                            test.get('test_id', ''),
                            category.upper(),
                            test.get('test_name', ''),
                            test.get('description', ''),
                            steps_str,
                            test.get('expected_result', ''),
                            test.get('priority', 'Medium'),
                            test.get('field', '') or test.get('element', '') or '',
                            'Not Executed',
                            '',
                            ''
                        ])
            
            return csv_file
        
        # Build XLSX report
        os.makedirs('reports', exist_ok=True)
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Summary sheet
        ws_summary = wb.create_sheet('Summary', 0)
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Summary header
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'] = 'Test Generation Summary'
        ws_summary['A1'].font = title_font
        ws_summary['A1'].alignment = center_align
        
        summary_rows = [
            ('Website URL', self.website_url),
            ('Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('', ''),
            ('Test Category', 'Count'),
        ]
        
        # Add category counts
        total = 0
        for category in ['positive', 'negative', 'ui', 'functional', 'workflow', 'edge_case', 'multi_level']:
            count = len(self.test_cases.get(category, []))
            total += count
            if count > 0:
                summary_rows.append((category.upper(), count))
        
        summary_rows.append(('TOTAL TESTS', total))
        
        # Write summary
        for r_idx, (k, v) in enumerate(summary_rows, start=2):
            cell_k = ws_summary.cell(row=r_idx, column=1, value=k)
            cell_v = ws_summary.cell(row=r_idx, column=2, value=v)
            if k and k != '':
                cell_k.font = header_font
            if r_idx == 2:  # Website URL
                cell_k.font = Font(bold=True)
                cell_v.font = Font(bold=True)
        
        # Adjust summary column widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50
        
        # All Test Cases sheet
        ws = wb.create_sheet('All Test Cases')
        
        # Headers
        headers = [
            'Test ID', 'Test Category', 'Test Name', 'Description (English)', 
            'Test Steps', 'Expected Result', 'Priority', 'Field/Element', 'Status', 'Result', 'Error Message'
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_color = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = header_font_color
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Write all test cases
        row = 2
        for category, test_list in self.test_cases.items():
            for test in test_list:
                # Format steps as readable text
                steps = test.get('steps', [])
                if isinstance(steps, list):
                    steps_str = '\n'.join([f"{idx+1}. {step}" for idx, step in enumerate(steps)]) if steps else 'N/A'
                else:
                    steps_str = str(steps) if steps else 'N/A'
                
                # Get description in English
                description = test.get('description', '')
                if not description:
                    description = test.get('test_name', '')
                
                # Get field/element name
                field_element = test.get('field', '') or test.get('element', '') or test.get('link_name', '') or ''
                
                ws.cell(row=row, column=1, value=test.get('test_id', '')).border = border
                ws.cell(row=row, column=2, value=category.upper()).border = border
                ws.cell(row=row, column=3, value=test.get('test_name', '')).border = border
                ws.cell(row=row, column=4, value=description).alignment = left_align
                ws.cell(row=row, column=4).border = border
                ws.cell(row=row, column=5, value=steps_str).alignment = left_align
                ws.cell(row=row, column=5).border = border
                ws.cell(row=row, column=6, value=test.get('expected_result', '')).alignment = left_align
                ws.cell(row=row, column=6).border = border
                ws.cell(row=row, column=7, value=test.get('priority', 'Medium')).border = border
                ws.cell(row=row, column=8, value=field_element).border = border
                ws.cell(row=row, column=9, value='Not Executed').border = border
                ws.cell(row=row, column=10, value='').border = border
                ws.cell(row=row, column=11, value='').alignment = left_align
                ws.cell(row=row, column=11).border = border
                
                row += 1
        
        # Auto-size columns
        column_widths = {
            'A': 15,  # Test ID
            'B': 12,  # Category
            'C': 30,  # Test Name
            'D': 40,  # Description
            'E': 50,  # Steps
            'F': 35,  # Expected Result
            'G': 10,  # Priority
            'H': 20,  # Field/Element
            'I': 12,  # Status
            'J': 10,  # Result
            'K': 40   # Error Message
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(xlsx_path)
        return xlsx_file
    
    def close(self):
        """Close browser"""
        try:
            if self.context:
                self.context.close()
            if self.browser:
                self.browser.close()
            if self.playwright:
                self.playwright.stop()
        except:
            pass
